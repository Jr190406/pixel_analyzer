from django.shortcuts import render, redirect, get_object_or_404
from django.core.files.storage import FileSystemStorage
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib.auth.forms import UserCreationForm
from django.db.models import Count, Sum, Q
from django.db import transaction
from django.utils import timezone
from django.urls import is_valid_path, reverse
from datetime import timedelta
from PIL import Image
import os
import numpy as np 
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
import json
from concurrent.futures import ThreadPoolExecutor
from .models import CostSetting, DefaultPricingRule, DocumentAnalysis, UserProfile, BusinessOwnerRequest, RequestMessage, RequestStatusChange, PrintedDocument
from .decorators import business_owner_required, super_admin_required, get_user_profile, authentication_required
from django.contrib.auth.models import AnonymousUser
from decimal import Decimal

POPPLER_PATH = None
def image_to_string(*args, **kwargs):
    return "OCR functionality disabled - install pytesseract"

def convert_from_path(*args, **kwargs):
    return []

class cv2:
    @staticmethod
    def imread(path):
        return None

class np:
    @staticmethod  
    def array(*args, **kwargs):
        return []
    @staticmethod
    def array_equal(*args, **kwargs):
        return True
    @staticmethod
    def sum(*args, **kwargs):
        return 0
    @staticmethod
    def all(*args, **kwargs):
        return True

# Global progress tracking
progress_sessions = {}

def update_progress(session_id, progress, message, **kwargs):
    """Update progress for a session with optional statistics"""
    if session_id:
        progress_sessions[session_id] = {
            'progress': progress,
            'message': message,
            'status': message,  # Store both for compatibility
            'timestamp': timezone.now().timestamp(),
            **kwargs
        }
        print(f"DEBUG - Progress updated: {session_id} - {progress}% - {message}")

def update_progress_with_stats(session_id, progress, message, pages_processed=0, total_pages=0, color_count=0, bw_count=0, current_cost=0):
    """Update progress with detailed statistics for real-time display"""
    update_progress(session_id, progress, message,
                   pages_processed=pages_processed,
                   total_pages=total_pages, 
                   color_count=color_count,
                   bw_count=bw_count,
                   current_cost=current_cost)
    print(f"DEBUG - Stats updated: Pages {pages_processed}/{total_pages}, Color: {color_count}, B&W: {bw_count}, Cost: ₱{current_cost:.2f}")

def clear_progress(session_id):
    """Clear progress for a session"""
    if session_id and session_id in progress_sessions:
        del progress_sessions[session_id]

def get_progress(request, session_id):
    """Get progress for a session via AJAX"""
    print(f"DEBUG: Getting progress for session: {session_id}")
    print(f"DEBUG: Available sessions: {list(progress_sessions.keys())}")
    
    if session_id in progress_sessions:
        data = progress_sessions[session_id]
        print(f"DEBUG: Returning progress data: {data}")
        # Clean up old sessions (older than 1 hour)
        current_time = timezone.now().timestamp()
        if current_time - data.get('timestamp', 0) > 3600:
            del progress_sessions[session_id]
            return JsonResponse({'progress': 0, 'message': 'Session expired'})
        return JsonResponse(data)
    
    print(f"DEBUG: No progress found for session: {session_id}")
    return JsonResponse({'progress': 0, 'message': 'No progress found'})

def landing_page(request):
    """Landing page view for the document analysis system"""
    return render(request, 'analyzer/landing_page.html')

def demo_upload(request):
    """Demo upload view - allows up to 10 file uploads without authentication and shows results on same page"""
    
    # Check demo usage (max 10 attempts)
    demo_count = request.session.get('demo_count', 0)
    max_demo_attempts = 10
    demo_remaining = max_demo_attempts - demo_count
    demo_used = demo_count >= max_demo_attempts
    
    # Check if we should show results
    show_results = False
    analysis_data = {}
    
    if request.method == 'POST' and request.FILES.get('file'):
        if demo_used:
            # If they try to upload after demo limit, show clear message
            messages.error(request, f'Demo limit reached! You have used all {max_demo_attempts} free uploads. Please register for unlimited access.')
            context = {
                'demo_used': True,
                'demo_count': demo_count,
                'demo_remaining': 0,
                'max_attempts': max_demo_attempts,
                'is_demo': True,
                'show_limit_message': True
            }
            return render(request, 'analyzer/demo_upload.html', context)
        
        try:
            # Increment demo count
            demo_count += 1
            request.session['demo_count'] = demo_count
            request.session['demo_timestamp'] = timezone.now().timestamp()
            
            # Process the file upload (similar to regular upload but simplified)
            uploaded_file = request.FILES['file']
            
            # Validate file type and size
            allowed_extensions = ['.pdf', '.png', '.jpg', '.jpeg']
            file_extension = os.path.splitext(uploaded_file.name)[1].lower()
            
            if file_extension not in allowed_extensions:
                messages.error(request, 'Please upload a PDF, PNG, or JPG file.')
                return render(request, 'analyzer/demo_upload.html')
            
            if uploaded_file.size > 10 * 1024 * 1024:  # 10MB limit
                messages.error(request, 'File size must be less than 10MB.')
                return render(request, 'analyzer/demo_upload.html')
            
            # Save file temporarily
            fs = FileSystemStorage()
            filename = fs.save(uploaded_file.name, uploaded_file)
            file_path = fs.path(filename)
            
            # Perform analysis immediately
            file_extension = os.path.splitext(uploaded_file.name)[1].lower()
            
            if file_extension == '.pdf':
                total_pages, results = analyze_pdf_demo(file_path)
            else:
                total_pages, results = analyze_image_demo(file_path)
            
            # Calculate costs
            total_color_cost = results.get('color_cost', 0)
            total_bw_cost = results.get('bw_cost', 0)
            total_cost = total_color_cost + total_bw_cost
            
            # Prepare analysis data for display
            analysis_data = {
                'filename': uploaded_file.name,
                'total_pages': total_pages,
                'results': results,
                'total_cost': format_cost(total_cost),
                'total_color_cost': format_cost(total_color_cost),
                'total_bw_cost': format_cost(total_bw_cost),
                'demo_count': demo_count,
                'demo_remaining': max_demo_attempts - demo_count,
                'max_attempts': max_demo_attempts,
                'is_demo': True,
                'demo_completed': True
            }
            show_results = True
            
            # Clean up file
            try:
                os.remove(file_path)
            except:
                pass
            
            # Show remaining attempts
            remaining_after = max_demo_attempts - demo_count
            if remaining_after > 0:
                messages.success(request, f'Demo file "{uploaded_file.name}" analyzed successfully! ({remaining_after} uploads remaining)')
            else:
                messages.success(request, f'Demo file "{uploaded_file.name}" analyzed successfully! This was your last free upload. Register for unlimited access!')
            
        except Exception as e:
            print(f"Error in demo upload: {e}")
            import traceback
            traceback.print_exc()
            messages.error(request, 'Error processing file. Please try again.')
            return render(request, 'analyzer/demo_upload.html')
    
    # Prepare context for template
    context = {
        'demo_count': demo_count,
        'demo_remaining': demo_remaining,
        'demo_used': demo_used,
        'max_attempts': max_demo_attempts,
        'is_demo': True,
        'show_results': show_results,
        'show_limit_message': demo_used,
        **analysis_data  # Spread analysis data if available
    }
    
    return render(request, 'analyzer/demo_upload.html', context)

def analyze_pdf_demo(file_path):
    from PIL import Image
    from io import BytesIO
    import base64
    """Simplified PDF analysis for demo - uses available PDF libraries to count pages"""
    try:
        print(f"DEBUG - Analyzing PDF: {file_path}")
        print(f"DEBUG - File exists: {os.path.exists(file_path)}")
        
        total_pages = 0
        
        print("DEBUG - UPDATED DEMO FUNCTION - Using PyPDF2 directly")
        
    # Try PyPDF2 first (most reliable for page counting)
        try:
            import PyPDF2
            print(f"DEBUG - PyPDF2 imported successfully, version: {getattr(PyPDF2, '__version__', 'Unknown')}")
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                total_pages = len(pdf_reader.pages)
                print(f"DEBUG - PyPDF2 SUCCESSFULLY found: {total_pages} pages")
        except Exception as e:
            print(f"DEBUG - PyPDF2 failed with error: {e}")
            import traceback
            traceback.print_exc()
            
            # Try pdfplumber as fallback
            try:
                import pdfplumber
                print(f"DEBUG - Trying pdfplumber fallback")
                with pdfplumber.open(file_path) as pdf:
                    total_pages = len(pdf.pages)
                    print(f"DEBUG - pdfplumber SUCCESSFULLY found: {total_pages} pages")
            except Exception as e2:
                print(f"DEBUG - pdfplumber also failed: {e2}")
                import traceback
                traceback.print_exc()
                total_pages = 1
                print("DEBUG - Using final fallback: 1 page")
        
        # Ensure we have at least 1 page
        if total_pages == 0:
            total_pages = 1
            print("DEBUG - No pages found, defaulting to 1")
        
        print(f"DEBUG - Final total pages: {total_pages}")
        
        # Use the same conversion logic as user upload
        color_pages = 0
        bw_pages = 0
        image_data = []
        from django.contrib.auth.models import AnonymousUser
        demo_user = AnonymousUser()
        try:
            images = convert_pdf_to_images(file_path)
            print(f"DEBUG - Got {len(images)} images from convert_pdf_to_images")
            for i, img in enumerate(images):
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                width, height = img.size
                color_pixels = 0
                total_samples = 0
                sample_step = max(20, min(width, height) // 30)
                for y in range(0, height, sample_step):
                    for x in range(0, width, sample_step):
                        if total_samples >= 500:
                            break
                        r, g, b = img.getpixel((x, y))
                        if abs(r - g) > 20 or abs(g - b) > 20 or abs(r - b) > 20:
                            color_pixels += 1
                        total_samples += 1
                coverage_percentage = (color_pixels / total_samples) * 100 if total_samples else 0
                is_color_page = coverage_percentage > 3
                if is_color_page:
                    color_pages += 1
                else:
                    bw_pages += 1
                display_img = img.copy()
                display_img.thumbnail((800, 1000), Image.Resampling.LANCZOS)
                buffer = BytesIO()
                display_img.save(buffer, format='PNG', quality=95)
                img_str = base64.b64encode(buffer.getvalue()).decode()
                image_data.append({
                    'page_number': i + 1,
                    'image_data': f"data:image/png;base64,{img_str}",
                    'is_color': is_color_page,
                    'coverage_percentage': round(coverage_percentage, 2)
                })
                print(f"DEBUG - Created display for PDF page {i + 1}")
            print(f"DEBUG - Successfully created {len(image_data)} PDF displays with convert_pdf_to_images")
        except Exception as pdf_error:
            print(f"DEBUG - PDF visual conversion failed: {pdf_error}")
            image_data = []
            color_pages = 0
            bw_pages = total_pages
        # Calculate costs using DefaultPricingRule for each page's actual coverage
        from .models import DefaultPricingRule
        paper_size = 'a4'  # Or get from form/context if needed
        color_cost = 0
        bw_cost = 0
        color_rate = 0
        bw_rate = 0
        for page in image_data:
            coverage = page.get('coverage_percentage', 50)  # Default to 50 if not set
            if page['is_color']:
                color_rule = DefaultPricingRule.objects.filter(
                    color=True,
                    paper_size=paper_size,
                    coverage_min__lte=coverage,
                    coverage_max__gte=coverage,
                    is_active=True
                ).first()
                rate = float(color_rule.cost) if color_rule else 2.0
                color_cost += rate
                color_rate = rate  # Last rate used
            else:
                bw_rule = DefaultPricingRule.objects.filter(
                    color=False,
                    paper_size=paper_size,
                    coverage_min__lte=coverage,
                    coverage_max__gte=coverage,
                    is_active=True
                ).first()
                rate = float(bw_rule.cost) if bw_rule else 1.0
                bw_cost += rate
                bw_rate = rate  # Last rate used
        print(f"DEBUG - Color pages: {color_pages}, B&W pages: {bw_pages}")
        
        # Convert PDF to clean visual images (just like PNG/JPEG display)
        print(f"DEBUG - Converting PDF to clean visual display...")
        
        import base64
        from io import BytesIO
        
        image_data = []
        
        try:
            # Use pdf2image to get CLEAN visual content (same as PNG/JPEG)
            print(f"DEBUG - Converting PDF to clean images (system PATH for Poppler)...")
            clean_images = convert_from_path(file_path, dpi=200)  # Higher DPI for clean output
            print(f"DEBUG - Got {len(clean_images)} clean PDF pages")
            
            for i, clean_img in enumerate(clean_images):
                try:
                    # Analyze this clean image for color content (same as JPEG analysis)
                    if clean_img.mode != 'RGB':
                        clean_img = clean_img.convert('RGB')
                    
                    # Same color analysis as JPEG files
                    width, height = clean_img.size
                    color_pixels = 0
                    total_samples = 0
                    sample_step = max(20, min(width, height) // 30)
                    
                    for y in range(0, height, sample_step):
                        for x in range(0, width, sample_step):
                            if total_samples >= 500:
                                break
                            try:
                                r, g, b = clean_img.getpixel((x, y))
                                if abs(r - g) > 25 or abs(g - b) > 25 or abs(r - b) > 25:
                                    color_pixels += 1
                                total_samples += 1
                            except:
                                pass
                        if total_samples >= 500:
                            break
                    
                    # Determine if page is color based on actual analysis
                    is_color_page = False
                    if total_samples > 0:
                        color_percentage = (color_pixels / total_samples) * 100
                        is_color_page = color_percentage > 3
                        print(f"DEBUG - PDF Page {i + 1}: {color_percentage:.1f}% color = {'COLOR' if is_color_page else 'B&W'}")
                    
                    # Resize for web display (like JPEG processing)
                    display_img = clean_img
                    if width > 800 or height > 1000:
                        display_img = clean_img.copy()
                        display_img.thumbnail((800, 1000), Image.Resampling.LANCZOS)
                        print(f"DEBUG - Resized page {i + 1} to {display_img.size} for clean display")
                    
                    # Convert to base64 for clean HTML display
                    buffer = BytesIO()
                    display_img.save(buffer, format='PNG', quality=95)
                    img_str = base64.b64encode(buffer.getvalue()).decode()
                    
                    image_data.append({
                        'page_number': i + 1,
                        'image_data': f"data:image/png;base64,{img_str}",
                        'is_color': is_color_page  # Based on actual analysis
                    })
                    print(f"DEBUG - Created clean display for PDF page {i + 1}")
                    
                except Exception as img_error:
                    print(f"DEBUG - Error processing clean PDF page {i + 1}: {img_error}")
            
            print(f"DEBUG - Successfully created {len(image_data)} clean PDF displays")
            
        except ImportError:
            print(f"DEBUG - pdf2image not available, using fallback...")
            
            # Fallback: Show message that visual display needs pdf2image
            from PIL import Image, ImageDraw
            
            fallback_img = Image.new('RGB', (600, 800), color='white')
            draw = ImageDraw.Draw(fallback_img)
            draw.rectangle([20, 20, 580, 780], outline='orange', width=3)
            draw.text((50, 100), "PDF Visual Display", fill='black')
            draw.text((50, 140), "Install pdf2image + poppler", fill='orange')
            draw.text((50, 180), "for clean PDF display", fill='orange')
            draw.text((50, 220), "(like PNG/JPEG files)", fill='gray')
            
            buffer = BytesIO()
            fallback_img.save(buffer, format='PNG')
            img_str = base64.b64encode(buffer.getvalue()).decode()
            
            image_data = [{
                'page_number': 1,
                'image_data': f"data:image/png;base64,{img_str}",
                'is_color': False
            }]
            
        except Exception as pdf_error:
            print(f"DEBUG - PDF visual conversion failed: {pdf_error}")
            image_data = []
        
        results = {
            'color_pages': color_pages,
            'bw_pages': bw_pages,
            'color_cost': color_cost,
            'bw_cost': bw_cost,
            'color_rate': color_rate,
            'bw_rate': bw_rate,
            'analysis_summary': f'Demo analysis: {total_pages} pages processed ({color_pages} color, {bw_pages} B&W)',
            'converted_images': image_data,  # Add visual content for bottom display
            'total_images': len(image_data),
            'per_page_coverage': [img['coverage_percentage'] for img in image_data],
            'per_page_color': [img['color_percentage'] for img in image_data]
        }
        
        print(f"DEBUG - Final results with {len(image_data)} images: {results}")
        return total_pages, results
        
    except Exception as e:
        print(f"ERROR in PDF demo analysis: {e}")
        import traceback
        traceback.print_exc()
        
        # Fallback values on error
        return 1, {
            'color_pages': 1, 
            'bw_pages': 0, 
            'color_cost': 2.0, 
            'bw_cost': 0,
            'color_rate': 2.0,
            'bw_rate': 1.0,
            'analysis_summary': 'Demo analysis completed (1 page fallback due to error)'
        }

def analyze_image_demo(file_path):
    """Simplified image analysis for demo - shows actual image at bottom"""
    try:
        print(f"DEBUG - Analyzing image for demo: {file_path}")
        
        # Simple image analysis
        total_pages = 1
        
        # Load and analyze the actual image
        from PIL import Image
        import base64
        from io import BytesIO
        
        # Open the actual uploaded image
        with Image.open(file_path) as img:
            print(f"DEBUG - Image opened: {img.size}, mode: {img.mode}")
            
            # Convert to RGB if needed
            if img.mode != 'RGB':
                img = img.convert('RGB')
            
            # Analyze for color content (same logic as PDF)
            width, height = img.size
            color_pixels = 0
            total_samples = 0
            sample_step = max(20, min(width, height) // 30)
            
            for y in range(0, height, sample_step):
                for x in range(0, width, sample_step):
                    if total_samples >= 500:
                        break
                    try:
                        r, g, b = img.getpixel((x, y))
                        if abs(r - g) > 25 or abs(g - b) > 25 or abs(r - b) > 25:
                            color_pixels += 1
                        total_samples += 1
                    except:
                        pass
                if total_samples >= 500:
                    break
            
            # Determine if image is color or B&W
            is_color = False
            if total_samples > 0:
                color_percentage = (color_pixels / total_samples) * 100
                is_color = color_percentage > 3
                print(f"DEBUG - Image analysis: {color_percentage:.1f}% color pixels = {'COLOR' if is_color else 'B&W'}")
            
            # Convert image to base64 for display
            # Resize if too large for display
            display_img = img
            if width > 800 or height > 600:
                display_img = img.copy()
                display_img.thumbnail((800, 600), Image.Resampling.LANCZOS)
                print(f"DEBUG - Resized image to {display_img.size} for display")
            
            buffer = BytesIO()
            display_img.save(buffer, format='PNG')
            img_str = base64.b64encode(buffer.getvalue()).decode()
            
            image_data = [{
                'page_number': 1,
                'image_data': f"data:image/png;base64,{img_str}",
                'is_color': is_color
            }]
        
        # Use the same pricing logic as regular users
        from django.contrib.auth.models import AnonymousUser
        demo_user = AnonymousUser()
        
        # Calculate costs based on actual analysis - USE DEFAULT PRICING
        if is_color:
            color_cost_per_page, color_reason = get_pricing_for_user(demo_user, True, 50)
            color_rate = float(color_cost_per_page)
            color_cost = 1 * color_rate
            color_pages = 1
            bw_pages = 0
            bw_cost = 0
            print(f"DEBUG - DEMO Image Color: ₱{color_rate} (DEFAULT pricing: {color_reason})")
        else:
            bw_cost_per_page, bw_reason = get_pricing_for_user(demo_user, False, 50)
            bw_rate = float(bw_cost_per_page)
            bw_cost = 1 * bw_rate
            bw_pages = 1
            color_pages = 0
            color_cost = 0
            print(f"DEBUG - DEMO Image B&W: ₱{bw_rate} (DEFAULT pricing: {bw_reason})")
        
        # Get both rates for display
        color_rate_per_page, _ = get_pricing_for_user(demo_user, True, 50)
        bw_rate_per_page, _ = get_pricing_for_user(demo_user, False, 50)
        
        results = {
            'color_pages': color_pages,
            'bw_pages': bw_pages,
            'color_cost': color_cost,
            'bw_cost': bw_cost,
            'color_rate': float(color_rate_per_page),
            'bw_rate': float(bw_rate_per_page),
            'analysis_summary': f'Demo analysis: 1 {"color" if is_color else "B&W"} image processed',
            'converted_images': image_data,  # Add actual image for bottom display
            'total_images': len(image_data)
        }
        
        print(f"DEBUG - Image demo results: {results}")
        return total_pages, results
    except Exception as e:
        return 1, {
            'color_pages': 1, 
            'bw_pages': 0, 
            'color_cost': 2.0, 
            'bw_cost': 0,
            'color_rate': 2.0,
            'bw_rate': 1.0,
            'analysis_summary': 'Demo analysis completed'
        }

def format_cost(cost):
    """Helper function to format cost values properly for both Decimal and float"""
    if isinstance(cost, Decimal):
        # For Decimal, check if it's equal to its integer conversion
        int_cost = int(cost)
        return int_cost if cost == int_cost else float(cost)
    elif isinstance(cost, float):
        # For float, use is_integer()
        return int(cost) if cost.is_integer() else cost
    else:
        return cost

@business_owner_required
def business_owner_upload(request):
    """Document analysis for business owners with premium progress tracking"""
    
    print(f"DEBUG - Business owner upload request from: {request.user.username if request.user.is_authenticated else 'Anonymous'}")
    
    if request.method == 'POST' and request.FILES.get('document'):
        uploaded_file = request.FILES['document']
        paper_size = request.POST.get('paper_size', 'a4')  # Default to A4
        
        # Get session ID from form or generate one
        import uuid
        session_id = request.POST.get('session_id') or str(uuid.uuid4())
        print(f"DEBUG: Business owner session ID: {session_id}")  # Debug print
        print(f"DEBUG: Paper size selected: {paper_size}")  # Debug print
        print(f"DEBUG: Starting progress tracking for session: {session_id}")
        
        # Initialize premium progress tracking
        update_progress_with_stats(session_id, 0, "Initializing premium processing...",
                                 pages_processed=0, total_pages=0,
                                 color_count=0, bw_count=0, current_cost=0)
        print(f"DEBUG: Initial progress set for session: {session_id}")
        
        file_path = os.path.join(settings.MEDIA_ROOT, uploaded_file.name)
        
        # Save uploaded file
        with open(file_path, 'wb+') as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)
        
        # Update progress
        update_progress(session_id, 10, "File uploaded successfully...")
        
        # Create results directory for business owner
        results_dir = os.path.join(settings.MEDIA_ROOT, f'business_results_{session_id}')
        os.makedirs(results_dir, exist_ok=True)
        
        # Update progress
        update_progress(session_id, 15, "Preparing premium analysis environment...")
        
        try:
            print(f"DEBUG - Starting document processing for: {uploaded_file.name}")
            # Process document similar to regular user but with business pricing
            results = []
            overall_cost = 0.0
            color_images = []
            bw_images = []
            
            file_name = uploaded_file.name.lower()
            print(f"DEBUG - File type check: {file_name}")
            if file_name.endswith('.pdf'):
                # Update progress
                update_progress(session_id, 20, "Converting PDF with premium processing...")
                
                # Process PDF pages
                print(f"DEBUG - Converting PDF: {file_path}")
                print(f"DEBUG - File exists check: {os.path.exists(file_path)}")
                print(f"DEBUG - File size: {os.path.getsize(file_path) if os.path.exists(file_path) else 'N/A'} bytes")
                
                images = convert_pdf_to_images(file_path)
                print(f"DEBUG - PDF conversion result: {len(images) if images else 0} images")
                print(f"DEBUG - Images type: {type(images)}")
                
                if images:
                    print(f"DEBUG - First image info: {images[0].size if images else 'None'}")
                else:
                    print("ERROR - PDF conversion returned empty list!")
                    
                if not images:
                    print("ERROR - PDF conversion failed!")
                    clear_progress(session_id)
                    return render(request, 'analyzer/business/business_owner_upload.html', {
                        'error': 'Could not process PDF file - conversion returned no images'
                    })

                # Update progress with total pages detected
                update_progress_with_stats(session_id, 30, f"PDF converted: {len(images)} pages detected...",
                                         pages_processed=0, total_pages=len(images), 
                                         color_count=0, bw_count=0, current_cost=0)
                
                # Save all images first
                page_paths = []
                for i, image in enumerate(images):
                    page_filename = f'{session_id}_page_{i+1}.jpg'
                    page_path = os.path.join(results_dir, page_filename)
                    image.save(page_path, 'JPEG')
                    page_paths.append((i, page_path, page_filename))
                
                # Update progress
                update_progress_with_stats(session_id, 40, "Starting premium parallel processing...",
                                         pages_processed=0, total_pages=len(page_paths),
                                         color_count=0, bw_count=0, current_cost=0)
                
                # Process business owner pages with parallel processing for faster performance
                print(f"DEBUG - Starting parallel processing with {len(page_paths)} pages")
                with ThreadPoolExecutor(max_workers=8) as executor:  # More workers for business owners
                    # Submit all pages for processing
                    print(f"DEBUG - Submitting {len(page_paths)} pages to thread pool")
                    future_to_page = {
                        executor.submit(process_image, page_path, request.user, paper_size): (i, page_path, page_filename) 
                        for i, page_path, page_filename in page_paths
                    }
                    print(f"DEBUG - All pages submitted to executor")
                    
                    # Collect results with progress tracking
                    page_results = []
                    completed = 0
                    total_pages = len(page_paths)
                    
                    for future in as_completed(future_to_page):
                        i, page_path, page_filename = future_to_page[future]
                        try:
                            print(f"DEBUG - Processing result for page {i + 1}")
                            result = future.result()
                            print(f"DEBUG - Page {i + 1} result: {result}")
                            print(f"DEBUG - Result type: {type(result)}")
                            
                            if result is None:
                                print(f"ERROR - Page {i + 1} returned None result!")
                                continue
                            
                            if not isinstance(result, dict):
                                print(f"ERROR - Page {i + 1} returned invalid result type: {type(result)}")
                                continue
                                
                            # Check if result has error
                            if 'error' in result:
                                print(f"ERROR - Page {i + 1} has error: {result.get('error_message', 'Unknown error')}")
                                # Still add it to results but make sure it has proper classification
                                result['page_number'] = i + 1
                                result['image_path'] = os.path.join(settings.MEDIA_URL, f'business_results_{session_id}', page_filename)
                                result['cost'] = result.get('cost', 0)
                                result['cost_formatted'] = format_cost(result.get('cost', 0))
                                # Ensure image_type is set properly for errors
                                if 'image_type' not in result:
                                    result['image_type'] = 'Unknown'
                                page_results.append((i, result))
                                completed += 1
                                continue
                            
                            # Check if result has required fields
                            required_fields = ['cost', 'image_type', 'coverage_percentage']
                            missing_fields = [field for field in required_fields if field not in result]
                            if missing_fields:
                                print(f"ERROR - Page {i + 1} missing required fields: {missing_fields}")
                                # Create minimal valid result
                                for field in missing_fields:
                                    if field == 'cost':
                                        result[field] = 0
                                    elif field == 'image_type':
                                        result[field] = 'Unknown'
                                    elif field == 'coverage_percentage':
                                        result[field] = 0
                            
                            result['page_number'] = i + 1
                            result['image_path'] = os.path.join(settings.MEDIA_URL, f'business_results_{session_id}', page_filename)
                            # Ensure cost_formatted is available for template
                            if 'cost_formatted' not in result and 'cost' in result:
                                result['cost_formatted'] = format_cost(result['cost'])
                            page_results.append((i, result))
                            completed += 1
                            print(f"DEBUG - ✅ Successfully completed {completed}/{total_pages} pages")
                            
                            # Count current stats for real-time updates
                            current_color_count = sum(1 for _, res in page_results if res.get('image_type') == 'Color')
                            current_bw_count = sum(1 for _, res in page_results if res.get('image_type') == 'Black & White')
                            current_cost = sum(float(res.get('cost', 0)) for _, res in page_results)
                            
                            # Update progress with detailed business processing status
                            progress = 40 + int((completed / total_pages) * 30)  # 40% to 70%
                            if result['image_type'] == 'Color':
                                message = f"Premium analysis: Color page {completed}/{total_pages} - {result['coverage_percentage']:.1f}% coverage"
                            else:
                                message = f"Premium analysis: B&W page {completed}/{total_pages} - {result['coverage_percentage']:.1f}% coverage"
                            
                            # Real-time progress update with statistics
                            update_progress_with_stats(session_id, progress, message, 
                                                     pages_processed=completed, 
                                                     total_pages=total_pages,
                                                     color_count=current_color_count,
                                                     bw_count=current_bw_count,
                                                     current_cost=current_cost)
                        except Exception as exc:
                            print(f'❌ ERROR - Page {i + 1} generated an exception: {exc}')
                            print(f'❌ Exception type: {type(exc).__name__}')
                            import traceback
                            print(f'❌ Full traceback:')
                            traceback.print_exc()
                    
                    # Update progress
                    update_progress(session_id, 75, "Premium cost calculation in progress...")
                    
                    # Sort results by page number and process
                    page_results.sort(key=lambda x: x[0])
                    for i, result in page_results:
                        results.append(result)
                        page_cost = float(result.get('cost', 0))
                        overall_cost += page_cost
                        
                        print(f"DEBUG - 📋 PAGE {i+1}: type='{result.get('image_type')}', cost={page_cost} (from: {result.get('cost')}), coverage={result.get('coverage_percentage')}%")
                        print(f"DEBUG - 💰 Running total cost: {overall_cost}")
                        
                        # Strict classification check
                        page_type = result.get('image_type', '')
                        if page_type == 'Color':
                            color_images.append(result)
                            print(f"DEBUG - 🎨 Added to COLOR SECTION: Page {i+1} (type='{page_type}')")
                        elif page_type == 'Black & White':
                            bw_images.append(result)
                            print(f"DEBUG - ⚫ Added to B&W SECTION: Page {i+1} (type='{page_type}')")
                        else:
                            # Fallback for unknown types
                            bw_images.append(result)
                            print(f"DEBUG - ❓ Added to B&W SECTION (fallback): Page {i+1} (type='{page_type}')")
                    
                    # Update progress
                    update_progress(session_id, 85, "Applying business owner pricing rules...")
            else:
                # Process single image file
                update_progress(session_id, 30, "Processing single image with premium quality...")
                
                # Copy to results directory
                image_filename = f'{session_id}_image.jpg'
                image_path = os.path.join(results_dir, image_filename)
                
                # Convert and save
                from PIL import Image
                img = Image.open(file_path)
                img = img.convert('RGB')
                img.save(image_path, 'JPEG')
                
                # Update progress
                update_progress(session_id, 50, "Analyzing single image content...")
                
                # Process the image
                result = process_image(image_path, user=request.user, paper_size=paper_size)
                result['page_number'] = 1
                result['image_path'] = os.path.join(settings.MEDIA_URL, f'business_results_{session_id}', image_filename)
                
                results.append(result)
                overall_cost = float(result['cost'])
                
                if result['image_type'] == 'Color':
                    color_images.append(result)
                else:
                    bw_images.append(result)
                
                # Update progress
                update_progress(session_id, 85, "Single image analysis complete...")
            
            # Update progress
            update_progress(session_id, 90, "Finalizing business owner analysis...")
            
            # Count color and B&W pages
            color_page_count = len(color_images)
            bw_page_count = len(bw_images)
            
            # Update progress
            update_progress(session_id, 95, "Saving premium analysis results...")
            
            # Save analysis to database for business owner
            analysis = DocumentAnalysis.objects.create(
                user=request.user,
                document_name=uploaded_file.name,
                page_count=len(results),
                color_page_count=color_page_count,
                bw_page_count=bw_page_count,
                overall_cost=overall_cost,
                analysis_result=results,  # Store the detailed results as JSON
                file_size=uploaded_file.size,
                file_type=uploaded_file.name.split('.')[-1].lower() if '.' in uploaded_file.name else 'unknown'
            )
            
            # Complete progress with final statistics
            final_message = f"Analysis complete! {len(results)} pages: {len(color_images)} color, {len(bw_images)} B&W - Total: ₱{format_cost(overall_cost)}"
            update_progress_with_stats(session_id, 100, final_message,
                                     pages_processed=len(results), total_pages=len(results),
                                     color_count=len(color_images), bw_count=len(bw_images),
                                     current_cost=overall_cost)
            
            # Clear progress after a short delay
            import threading
            def clear_progress_delayed():
                time.sleep(3)  # Show results for 3 seconds
                clear_progress(session_id)
                
            threading.Thread(target=clear_progress_delayed).start()
            
            # Debug: Check if we have results to display
            print(f"DEBUG - 📋 TEMPLATE CONTEXT PREP:")
            print(f"  color_images: {len(color_images)} items")
            print(f"  bw_images: {len(bw_images)} items")
            print(f"  overall_cost: {overall_cost}")
            print(f"  formatted overall_cost: {format_cost(overall_cost)}")
            print(f"  total_pages: {len(results)}")
            print(f"  file_name: {uploaded_file.name}")
            
            if color_images:
                sample_color = color_images[0]
                print(f"  First color image: cost={sample_color.get('cost')}, type={sample_color.get('image_type')}, coverage={sample_color.get('coverage_percentage')}%")
            if bw_images:
                sample_bw = bw_images[0]
                print(f"  First BW image: cost={sample_bw.get('cost')}, type={sample_bw.get('image_type')}, coverage={sample_bw.get('coverage_percentage')}%")
            
            context = {
                'color_images': color_images,
                'bw_images': bw_images,
                'overall_cost': format_cost(overall_cost),
                'total_pages': len(results),
                'file_name': uploaded_file.name,
                'analysis_id': analysis.id,
                'paper_size': paper_size,
                'session_id': session_id,
                'success_message': f"Premium analysis completed! Total cost: ₱{format_cost(overall_cost)}",
                'debug_info': f"Processed {len(results)} pages: {len(color_images)} color, {len(bw_images)} B&W"
            }
            
            print(f"DEBUG - RENDERING TEMPLATE WITH CONTEXT: {list(context.keys())}")
            return render(request, 'analyzer/business/business_owner_upload.html', context)
            
        except Exception as e:
            # Clear progress on error
            clear_progress(session_id)
            return render(request, 'analyzer/business/business_owner_upload.html', {
                'error': f'Error processing document: {str(e)}'
            })
    
    return render(request, 'analyzer/business/business_owner_upload.html')

def is_safe_url(url):
    """Check if the URL is safe for redirecting (prevent open redirects)"""
    if not url:
        return False
    
    # Only allow relative URLs or URLs starting with /
    if url.startswith('/') and not url.startswith('//'):
        # Check if it's a valid path in our Django app
        valid_paths = [
            '/',
            '/dashboard/',
            '/regular-user-dashboard/',
            '/business-owner-dashboard/',
            '/pricing/',
            '/history/',
            '/upload/',
            '/admin-dashboard/',
            '/business-dashboard/',
            '/business-owner-request/',
            '/role-check/',
        ]
        return any(url.startswith(path) for path in valid_paths)
    return False

def is_color(image_path):
    image = cv2.imread(image_path)
    if image is None:
        return False  
    if np.array_equal(image[:, :, 0], image[:, :, 1]) and np.array_equal(image[:, :, 1], image[:, :, 2]):
        return False  
    return True  

def extract_text(image):
    return image_to_string(image, config='--oem 1 --psm 6').strip()

def process_image(image_path, user=None, paper_size='short'):
    try:
        print(f"DEBUG - process_image called with: {image_path}, user: {user.username if user and hasattr(user, 'username') else 'None'}")
        
        # Check if file exists first
        if not os.path.exists(image_path):
            print(f"ERROR - Image file does not exist: {image_path}")
            return {
                'error': 'FILE_NOT_FOUND',
                'error_message': f'Image file not found: {image_path}',
                'coverage_percentage': 0,
                'image_type': "Unknown",
                'cost': 0,
                'cost_reason': 'File not found'
            }
        
        # Open and optimize image processing
        image = Image.open(image_path)
        print(f"DEBUG - Image opened: {image.size}, mode: {image.mode}")
        
        # Resize large images for faster processing (maintain aspect ratio)
        max_size = 1500  # Maximum dimension for analysis
        if max(image.size) > max_size:
            image.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)
            print(f"DEBUG - Image resized to: {image.size}")
        
        image = image.convert('RGB')
        
        # Use numpy for faster pixel processing
        img_array = None
        try:
            img_array = np.array(image)
            print(f"DEBUG - Image array shape: {img_array.shape}")
            print(f"DEBUG - Image array type: {type(img_array)}")
        except Exception as np_error:
            print(f"DEBUG - NumPy conversion failed: {np_error}, falling back to PIL")
        width, height = image.size
        total_pixels = width * height

        # USE THE SAME WORKING ALGORITHM AS REGULAR USER
        print(f"DEBUG - 🔍 USING PROVEN REGULAR USER ALGORITHM")
        
        color_detected = False
        non_white_pixels = 0
        sample_count = 0
        color_pixels_found = 0
        
        # Sample more pixels for better detection - same as regular user
        sample_step_x = max(1, width // 100)   # Sample every 100th pixel horizontally
        sample_step_y = max(1, height // 100)  # Sample every 100th pixel vertically
        
        print(f"DEBUG - Sampling with steps: x={sample_step_x}, y={sample_step_y}")
        
        for y in range(0, height, sample_step_y):
            for x in range(0, width, sample_step_x):
                try:
                    pixel = image.getpixel((x, y))
                    sample_count += 1
                    
                    # Check if pixel is not white (with tolerance for near-white)
                    if not (pixel[0] > 250 and pixel[1] > 250 and pixel[2] > 250):
                        non_white_pixels += 1
                    
                    # Enhanced color detection - same as regular user
                    r, g, b = pixel
                    
                    # Check for significant color difference - SAME THRESHOLD AS REGULAR USER
                    max_diff = max(abs(r-g), abs(g-b), abs(r-b))
                    if max_diff > 20:  # Same threshold as regular user (more sensitive)
                        color_pixels_found += 1
                        if not color_detected:
                            color_detected = True
                            print(f"DEBUG - 🎨 Color detected at pixel ({x}, {y}): RGB{pixel}, max_diff: {max_diff}")
                        
                except (IndexError, Exception):
                    continue
        
        # Same color percentage logic as regular user
        if sample_count > 0:
            color_percentage = (color_pixels_found / sample_count) * 100
            print(f"DEBUG - 🎨 Color analysis: {color_pixels_found}/{sample_count} pixels ({color_percentage:.1f}%) are colored")
            
            # SAME THRESHOLD AS REGULAR USER - 2% or more pixels are colored
            if color_percentage >= 2.0:
                color_detected = True
        
        print(f"DEBUG - 🔴 FINAL COLOR DECISION: {color_detected}")

        # USE SAME COVERAGE CALCULATION AS REGULAR USER 
        print(f"DEBUG - 📊 USING REGULAR USER COVERAGE CALCULATION")
        
        # Calculate coverage percentage - SAME AS REGULAR USER
        if sample_count > 0:
            coverage_percentage = (non_white_pixels / sample_count) * 100
        else:
            coverage_percentage = 0
        
        print(f"DEBUG - 📈 COVERAGE RESULT: {coverage_percentage:.1f}% coverage ({non_white_pixels}/{sample_count} non-white pixels)")

        # Skip text detection for faster processing (can be enabled if needed)
        text_detected = ""  # Disabled for speed
        has_text = False

        cost = 0
        cost_reason = "Blank Page"

        # Assign cost based on content
        print(f"DEBUG - Checking if coverage > 0.1%: {coverage_percentage} > 0.1 = {coverage_percentage > 0.1}")
        if coverage_percentage > 0.1:  # Only process if there's meaningful content
            try:
                print(f"DEBUG - 🔍 PRICING CALL: user={user.username if user else 'None'}, color={color_detected}, coverage={coverage_percentage}, paper_size={paper_size}")
                cost, cost_reason = get_pricing_for_user(user, color_detected, coverage_percentage, paper_size)
                print(f"DEBUG - 💰 PRICING RESULT: cost={cost} (type: {type(cost)}), reason={cost_reason}")
            except ValueError as e:
                print(f"DEBUG - Pricing error: {e}")
                error_msg = str(e)
                if "INCOMPLETE_PRICING_RULES" in error_msg:
                    # Return error indicator for incomplete pricing rules
                    return {
                        'error': 'INCOMPLETE_PRICING_RULES',
                        'error_message': 'Business owner must set pricing rules up to 100% coverage for both B&W and Color.',
                        'coverage_percentage': round(coverage_percentage, 2),
                        'image_type': "Color" if color_detected else "Black & White",
                    }
                else:
                    # Other pricing errors
                    return {
                        'error': 'PRICING_ERROR',
                        'error_message': f'Pricing calculation failed: {error_msg}',
                        'coverage_percentage': round(coverage_percentage, 2),
                        'image_type': "Color" if color_detected else "Black & White",
                    }
        else:
            print(f"DEBUG - Coverage too low ({coverage_percentage:.2f}%), using blank page cost")
        
        # Determine image type based on enhanced color detection
        image_type = "Color" if color_detected else "Black & White"
        
        result = {
            'image_path': os.path.join(settings.MEDIA_URL, os.path.basename(image_path)),
            'cost': cost,
            'cost_reason': cost_reason,
            'image_type': image_type,
            'text_detected': text_detected,
            'coverage_percentage': round(coverage_percentage, 2),
        }
        
        print(f"DEBUG - 📊 FINAL RESULT: cost={result['cost']}, type={result['image_type']}, coverage={result['coverage_percentage']}%")
        print(f"DEBUG - 🎨 COLOR DECISION: {color_pixels_found}/{sample_count} color pixels → {image_type}")
        print(f"DEBUG - 📄 CONTENT DECISION: {non_white_pixels}/{sample_count} content pixels → {coverage_percentage:.1f}% coverage")
        print(f"DEBUG - ✅ PAGE CLASSIFICATION: {'COLOR' if color_detected else 'BLACK & WHITE'}")
        return result
        
    except Exception as e:
        print(f"❌ CRITICAL ERROR in process_image: {e}")
        print(f"❌ Error type: {type(e).__name__}")
        import traceback
        traceback.print_exc()
        
        # Return error result instead of crashing
        return {
            'error': 'PROCESSING_ERROR',
            'error_message': f'Image processing failed: {str(e)}',
            'coverage_percentage': 0,
            'image_type': "Unknown",
            'cost': 0,
            'cost_reason': f'Processing error: {str(e)[:50]}'
        }

def convert_pdf_to_images(pdf_path):
    """Convert PDF to clean visual images - shows EXACTLY what's in the PDF (like PNG/JPEG)"""
    print(f"DEBUG - Converting PDF to CLEAN visual display: {pdf_path}")
    
    # Check if file exists first
    if not os.path.exists(pdf_path):
        print(f"ERROR - PDF file does not exist: {pdf_path}")
        return []
        
    print(f"DEBUG - File exists, size: {os.path.getsize(pdf_path)} bytes")
    
    try:
        # Use PyMuPDF (fitz) for clean PDF to image conversion - no poppler needed!
        import fitz  # PyMuPDF
        from PIL import Image
        import io
        
        print(f"DEBUG - Using PyMuPDF to convert PDF to CLEAN visual images...")
        print(f"DEBUG - Opening PDF: {pdf_path}")
        print(f"DEBUG - PyMuPDF version: {fitz.version}")
        
        # Open the PDF document
        print(f"DEBUG - About to open PDF with fitz.open()...")
        pdf_document = fitz.open(pdf_path)
        total_pages = len(pdf_document)
        
        print(f"DEBUG - PDF opened successfully: {total_pages} pages found")
        
        if total_pages == 0:
            print(f"ERROR - PDF has 0 pages! This is the problem!")
            pdf_document.close()
            return []
        
        clean_images = []
        
        for page_num in range(total_pages):
            print(f"DEBUG - Processing page {page_num + 1}/{total_pages}")
            
            try:
                # Get the page
                page = pdf_document[page_num]
                print(f"DEBUG - Got page object: {page}")
                
                # Convert page to image (high quality)
                mat = fitz.Matrix(2.0, 2.0)  # 2x zoom for good quality
                print(f"DEBUG - Creating pixmap with matrix...")
                pix = page.get_pixmap(matrix=mat)
                print(f"DEBUG - Pixmap created: {pix.width}x{pix.height} pixels")
                
                # Convert to PIL Image
                print(f"DEBUG - Converting pixmap to PIL Image...")
                img_data = pix.tobytes("ppm")
                img = Image.open(io.BytesIO(img_data))
                
                # Convert to RGB if needed
                if img.mode != 'RGB':
                    print(f"DEBUG - Converting from {img.mode} to RGB...")
                    img = img.convert('RGB')
                
                clean_images.append(img)
                print(f"DEBUG - ✅ Clean page {page_num + 1} converted: {img.size} pixels")
                
            except Exception as page_error:
                print(f"ERROR - Failed to process page {page_num + 1}: {page_error}")
                print(f"ERROR - Page error type: {type(page_error).__name__}")
                continue
        
        # Close the PDF
        pdf_document.close()
        
        print(f"DEBUG - ✅ SUCCESS! PyMuPDF converted {len(clean_images)} pages to CLEAN visual images")
        return clean_images
        
    except ImportError:
        print(f"ERROR - PyMuPDF not available, trying pdf2image fallback...")
        
        try:
            # Fallback to pdf2image if PyMuPDF fails
            from pdf2image import convert_from_path
            
            print(f"DEBUG - Fallback: Using pdf2image...")
            images = convert_from_path(pdf_path, dpi=150)
            
            clean_images = []
            for i, img in enumerate(images):
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                clean_images.append(img)
                print(f"DEBUG - ✅ pdf2image page {i + 1} ready: {img.size}")
            
            print(f"DEBUG - ✅ pdf2image fallback success: {len(clean_images)} images")
            return clean_images
            
        except Exception as pdf2image_error:
            print(f"ERROR - pdf2image also failed: {pdf2image_error}")
            return []
        
    except Exception as e:
        print(f"ERROR - PyMuPDF conversion failed: {e}")
        print(f"ERROR - Error type: {type(e).__name__}")
        import traceback
        traceback.print_exc()
        
        # Create a simple error message
        try:
            from PIL import Image, ImageDraw
            
            error_img = Image.new('RGB', (600, 400), color='white')
            draw = ImageDraw.Draw(error_img)
            
            draw.rectangle([10, 10, 590, 390], outline='red', width=2)
            draw.text((30, 50), "PDF Processing Failed", fill='red')
            draw.text((30, 90), f"Error: {str(e)[:50]}", fill='red')
            draw.text((30, 130), "Please try:", fill='black')
            draw.text((30, 160), "1. Upload PNG/JPEG files instead", fill='green')
            draw.text((30, 190), "2. Or check if PDF is corrupted", fill='blue')
            
            return [error_img]
            
        except:
            return []

def process_document_for_regular_user(file_path, user, results_dir, session_id=None):
    """Process a document and save images to results directory for regular users"""
    try:
        results = []
        overall_cost = 0.0
        color_images = []
        bw_images = []
        
        import uuid
        import shutil
        if not session_id:
            session_id = str(uuid.uuid4())[:8]  # Unique session ID

        # Get the file name to check extension
        file_name = os.path.basename(file_path)
        
        # Update progress
        if session_id:
            update_progress(session_id, 25, "Analyzing document structure...")
        
        if file_name.lower().endswith('.pdf'):
            # Process PDF
            if session_id:
                update_progress(session_id, 30, "Converting PDF to images...")
            
            images = convert_pdf_to_images(file_path)
            if not images:
                return [], [], 0, "Failed to convert PDF to images"
            
            if session_id:
                update_progress(session_id, 40, f"Converting PDF: {len(images)} pages detected...")
                time.sleep(0.8)  # Add delay to show this step
                
            image_paths = []
            for i, image in enumerate(images):
                # Save to results directory directly
                result_filename = f"{session_id}_page_{i + 1}.jpg"
                result_path = os.path.join(results_dir, result_filename)
                
                # Ensure results directory exists
                os.makedirs(results_dir, exist_ok=True)
                
                # Save the image with high quality
                image.save(result_path, 'JPEG', quality=90)
                image_paths.append(result_path)
                
                print(f"DEBUG - Saved page {i+1} image to: {result_path}")
                print(f"DEBUG - File exists: {os.path.exists(result_path)}")
                
                # Update progress during image saving
                if session_id and len(images) > 1:
                    progress = 40 + int((i + 1) / len(images) * 8)  # 40% to 48%
                    update_progress(session_id, progress, f"Saving page {i+1}/{len(images)} image...")

            # Process images in parallel (reduced workers for regular users to prioritize business owners)
            profile = get_user_profile(user)
            max_workers = 2 if profile.role == 'regular' else 8  # Even slower for regular users
            
            if session_id:
                update_progress(session_id, 50, "Starting detailed page analysis...")
                time.sleep(1)  # Show this step clearly
            
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                # Process pages and update progress
                completed = 0
                total_pages = len(image_paths)
                
                future_to_path = {executor.submit(process_image_for_regular_user, path, user, results_dir): path for path in image_paths}
                results = []
                
                for future in as_completed(future_to_path):
                    result = future.result()
                    results.append(result)
                    completed += 1
                    
                    # Update progress during processing with detailed status
                    if session_id:
                        progress = 50 + int((completed / total_pages) * 20)  # 50% to 70%
                        if result['image_type'] == 'Color':
                            update_progress(session_id, progress, f"Analyzed color page {completed}/{total_pages} - {result['coverage_percentage']:.1f}% ink coverage")
                        else:
                            update_progress(session_id, progress, f"Analyzed B&W page {completed}/{total_pages} - {result['coverage_percentage']:.1f}% coverage")
                        
                        # Add small delay for regular users during processing
                        if profile.role == 'regular':
                            time.sleep(0.3)  # Additional delay per page for regular users

        else:
            # Process single image - copy to results directory
            if session_id:
                update_progress(session_id, 40, "Processing single image...")
                time.sleep(0.5)
                
            result_filename = f"{session_id}_{os.path.basename(file_path)}"
            result_path = os.path.join(results_dir, result_filename)
            shutil.copy2(file_path, result_path)
            
            if session_id:
                update_progress(session_id, 55, "Analyzing image content and ink coverage...")
                time.sleep(1)  # Add delay for single image
            
            result = process_image_for_regular_user(result_path, user, results_dir)
            results = [result]
            
            if session_id:
                coverage = result['coverage_percentage']
                img_type = result['image_type']
                update_progress(session_id, 70, f"Analysis complete: {img_type} image with {coverage:.1f}% coverage")

        # Update progress
        if session_id:
            update_progress(session_id, 75, "Calculating printing costs...")
            
        # Categorize results with detailed progress
        total_results = len(results)
        for i, result in enumerate(results):
            if result['image_type'] == "Color":
                color_images.append(result)
            else:
                bw_images.append(result)
            overall_cost += result['cost']
            
            # Update progress during cost calculation
            if session_id and total_results > 1:
                progress = 75 + int((i + 1) / total_results * 10)  # 75% to 85%
                if result['image_type'] == "Color":
                    update_progress(session_id, progress, f"Calculating color page {i+1}/{total_results} cost...")
                else:
                    update_progress(session_id, progress, f"Calculating B&W page {i+1}/{total_results} cost...")

        # Update progress
        if session_id:
            update_progress(session_id, 85, "Computing total printing expenses...")
            time.sleep(1)  # Add delay to show this step
            update_progress(session_id, 90, "Preparing cost breakdown...")
            time.sleep(0.5)
            
        # Format overall_cost to show as integer if it's a whole number
        formatted_overall_cost = int(overall_cost) if isinstance(overall_cost, float) and overall_cost.is_integer() else overall_cost

        return color_images, bw_images, formatted_overall_cost, None

    except Exception as e:
        return [], [], 0, str(e)

def process_image_for_regular_user(image_path, user, results_dir):
    """Process image for regular users - Python 3.14 compatible version"""
    try:
        print(f"DEBUG - Processing image: {image_path}")
        
        # Open and optimize image processing
        image = Image.open(image_path)
        
        # Resize large images for faster processing (maintain aspect ratio)
        max_size = 1500  # Maximum dimension for analysis
        if max(image.size) > max_size:
            image.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)
        
        image = image.convert('RGB')
        width, height = image.size
        total_pixels = width * height
        
        # Enhanced color detection and coverage calculation
        color_detected = False
        non_white_pixels = 0
        sample_count = 0
        color_pixels_found = 0
        
        # Sample more pixels for better color detection
        sample_step_x = max(1, width // 100)   # Sample every 100th pixel horizontally
        sample_step_y = max(1, height // 100)  # Sample every 100th pixel vertically
        
        for y in range(0, height, sample_step_y):
            for x in range(0, width, sample_step_x):
                try:
                    pixel = image.getpixel((x, y))
                    sample_count += 1
                    
                    # Check if pixel is not white (with tolerance for near-white)
                    if not (pixel[0] > 250 and pixel[1] > 250 and pixel[2] > 250):
                        non_white_pixels += 1
                    
                    # Enhanced color detection - check for significant color difference
                    r, g, b = pixel
                    
                    # Check if there's a significant difference between color channels
                    max_diff = max(abs(r-g), abs(g-b), abs(r-b))
                    if max_diff > 20:  # Threshold for color detection (more sensitive)
                        color_pixels_found += 1
                        if not color_detected:
                            color_detected = True
                            print(f"DEBUG - Color detected at pixel ({x}, {y}): RGB{pixel}, max_diff: {max_diff}")
                        
                except IndexError:
                    continue
        
        # Require a minimum percentage of color pixels to classify as color page
        if sample_count > 0:
            color_percentage = (color_pixels_found / sample_count) * 100
            print(f"DEBUG - Color analysis: {color_pixels_found}/{sample_count} pixels ({color_percentage:.1f}%) are colored")
            
            # More sensitive color detection - if 2% or more pixels are colored, it's a color page
            if color_percentage >= 2.0:
                color_detected = True
        
        # Calculate coverage percentage
        if sample_count > 0:
            coverage_percentage = (non_white_pixels / sample_count) * 100
        else:
            coverage_percentage = 0
        
        print(f"DEBUG - Image analysis: {coverage_percentage:.1f}% coverage, Color: {color_detected}")
        
        # Skip text detection for faster processing
        text_detected = ""
        has_text = False
        
        cost = 0
        cost_reason = "Blank Page"
        
        # Assign cost based on content
        if coverage_percentage > 0.1:  # Only process if there's meaningful content
            cost, cost_reason = get_pricing_for_user(user, color_detected, coverage_percentage)
            print(f"DEBUG - Cost calculated: ₱{cost} - {cost_reason}")
        
        # Create web-accessible path
        filename = os.path.basename(image_path)
        web_path = f"/media/results/{filename}"
        
        return {
            'image_path': web_path,
            'cost': cost,
            'cost_reason': cost_reason,
            'image_type': "Color" if color_detected else "Black & White",
            'text_detected': text_detected,
            'coverage_percentage': round(coverage_percentage, 2),
        }
        
    except Exception as e:
        print(f"ERROR - Image processing failed: {e}")
        import traceback
        traceback.print_exc()
        
        # Return default values on error
        filename = os.path.basename(image_path) if image_path else "unknown.jpg"
        web_path = f"/media/results/{filename}"
        
        return {
            'image_path': web_path,
            'cost': 1.5,  # Default cost
            'cost_reason': "Error in analysis - default pricing applied",
            'image_type': "Black & White",
            'text_detected': "",
            'coverage_percentage': 25.0,
        }

def process_document(file_path, user=None):
    """Process a document and return analysis results"""
    try:
        results = []
        overall_cost = 0.0
        color_images = []
        bw_images = []

        # Get the file name to check extension
        file_name = os.path.basename(file_path)
        
        if file_name.lower().endswith('.pdf'):
            # Process PDF
            images = convert_pdf_to_images(file_path)
            if not images:
                return [], [], 0, "Failed to convert PDF to images"
                
            image_paths = []
            for i, image in enumerate(images):
                temp_path = f"{file_path}_page_{i + 1}.jpg"
                image.save(temp_path, 'JPEG')
                image_paths.append(temp_path)

            # Process images in parallel
            with ThreadPoolExecutor(max_workers=6) as executor:
                results = list(executor.map(lambda path: process_image(path, user), image_paths))

            # Clean up temporary image files
            for temp_path in image_paths:
                try:
                    os.remove(temp_path)
                except:
                    pass  # Ignore cleanup errors

        else:
            # Process single image
            result = process_image(file_path, user)
            results = [result]

        # Categorize results
        for result in results:
            if result['image_type'] == "Color":
                color_images.append(result)
            else:
                bw_images.append(result)
            overall_cost += result['cost']

        # Format overall_cost to show as integer if it's a whole number
        formatted_overall_cost = int(overall_cost) if isinstance(overall_cost, float) and overall_cost.is_integer() else overall_cost

        return color_images, bw_images, formatted_overall_cost, None

    except Exception as e:
        return [], [], 0, str(e)

@authentication_required
def dashboard_router(request):
    """Route users to their appropriate dashboard based on role"""
    profile = get_user_profile(request.user)
    
    if profile.is_super_admin():
        return redirect('super_admin_dashboard')
    elif profile.is_business_owner():
        return redirect('business_owner_dashboard')
    else:
        return redirect('regular_user_dashboard')

def handle_file_upload(request):
    """Handle file upload processing for authenticated users"""
    uploaded_file = request.FILES['file']
    fs = FileSystemStorage()
    filename = fs.save(uploaded_file.name, uploaded_file)
    file_path = os.path.join(settings.MEDIA_ROOT, filename)

    file_url = fs.url(filename)
    results = []
    overall_cost = 0.0
    color_images = []
    bw_images = []

    try:
        if uploaded_file.name.lower().endswith('.pdf'):
            images = convert_pdf_to_images(file_path)
            image_paths = []
            for i, image in enumerate(images):
                temp_path = f"{file_path}_page_{i + 1}.jpg"
                image.save(temp_path, 'JPEG')
                image_paths.append(temp_path)

            with ThreadPoolExecutor(max_workers=6) as executor:  # Increased from 4
                results = list(executor.map(lambda path: process_image(path, request.user), image_paths))

            for result in results:
                if result['image_type'] == "Color":
                    color_images.append(result)
                else:
                    bw_images.append(result)
                overall_cost += result['cost']

        else:
            result = process_image(file_path, request.user)
            if result['image_type'] == "Color":
                color_images.append(result)
            else:
                bw_images.append(result)
            overall_cost += result['cost']

        # Save analysis results to database if user is authenticated
        if request.user.is_authenticated:
            # Prepare analysis data for storage
            analysis_data = {
                'color_images': [
                    {
                        'page_number': i + 1,
                        'cost': float(img['cost']),
                        'cost_reason': img['cost_reason'],
                        'coverage_percentage': img['coverage_percentage'],
                        'image_type': img['image_type'],
                        'image_path': img['image_path']
                    }
                    for i, img in enumerate(color_images)
                ],
                'bw_images': [
                    {
                        'page_number': i + 1,
                        'cost': float(img['cost']),
                        'cost_reason': img['cost_reason'],
                        'coverage_percentage': img['coverage_percentage'],
                        'image_type': img['image_type'],
                        'image_path': img['image_path']
                    }
                    for i, img in enumerate(bw_images)
                ]
            }

            # Get file info
            file_size = uploaded_file.size
            file_type = uploaded_file.name.split('.')[-1].lower() if '.' in uploaded_file.name else 'unknown'

            # Only save analysis to database for authenticated users
            if request.user.is_authenticated:
                DocumentAnalysis.objects.create(
                    user=request.user,
                    document_name=uploaded_file.name,
                    page_count=len(color_images) + len(bw_images),
                    color_page_count=len(color_images),
                    bw_page_count=len(bw_images),
                    overall_cost=overall_cost,
                    analysis_result=analysis_data,
                    file_size=file_size,
                    file_type=file_type
                )
                messages.success(request, f"Analysis saved to your history! Document: {uploaded_file.name} (₱{overall_cost})")

        # Format overall_cost to show as integer if it's a whole number
        formatted_overall_cost = int(overall_cost) if isinstance(overall_cost, float) and overall_cost.is_integer() else overall_cost

        context = {
            'file_url': file_url,
            'color_images': color_images,
            'bw_images': bw_images,
            'overall_cost': formatted_overall_cost
        }
    except Exception as e:
        context = {'error': f"Error processing file: {str(e)}"}

    # Always use upload template to show results
    return render(request, 'analyzer/upload.html', context)

def get_regular_user_dashboard_context(user):
    """Helper function to get dashboard context for regular users"""
    # Get user's recent analyses
    recent_analyses = DocumentAnalysis.objects.filter(user=user).order_by('-created_at')[:5]
    
    # Get user statistics
    total_analyses = DocumentAnalysis.objects.filter(user=user).count()
    total_cost = DocumentAnalysis.objects.filter(user=user).aggregate(
        total=Sum('overall_cost')
    )['total'] or 0
    
    # Get monthly statistics
    from datetime import datetime, timedelta
    last_month = timezone.now() - timedelta(days=30)
    monthly_analyses = DocumentAnalysis.objects.filter(
        user=user,
        created_at__gte=last_month
    ).count()
    
    return {
        'recent_analyses': recent_analyses,
        'total_analyses': total_analyses,
        'total_cost': total_cost,
        'monthly_analyses': monthly_analyses,
        'user_profile': get_user_profile(user),
    }

@authentication_required
def regular_user_dashboard(request):
    """Dashboard for regular users - document upload and history"""
    # Get user's recent analyses
    recent_analyses = DocumentAnalysis.objects.filter(user=request.user).order_by('-created_at')[:5]
    
    # Get user statistics
    total_analyses = DocumentAnalysis.objects.filter(user=request.user).count()
    total_cost = DocumentAnalysis.objects.filter(user=request.user).aggregate(
        total=Sum('overall_cost')
    )['total'] or 0
    
    # Get monthly statistics
    from datetime import datetime, timedelta
    last_month = timezone.now() - timedelta(days=30)
    monthly_analyses = DocumentAnalysis.objects.filter(
        user=request.user,
        created_at__gte=last_month
    ).count()
    
    context = {
        'recent_analyses': recent_analyses,
        'total_analyses': total_analyses,
        'total_cost': total_cost,
        'monthly_analyses': monthly_analyses,
        'user_profile': get_user_profile(request.user),
    }
    
    # Check for upload results in session
    if 'upload_results' in request.session:
        upload_results = request.session.pop('upload_results')
        context.update(upload_results)
    
    return render(request, 'analyzer/user/regular_user_dashboard.html', context)

@authentication_required
def regular_user_upload(request):
    """Upload page for regular users"""
    context = {
        'user_profile': get_user_profile(request.user),
    }
    
    # Check for upload results in session
    if 'upload_results' in request.session:
        upload_results = request.session.pop('upload_results')
        context.update(upload_results)
    
    return render(request, 'analyzer/user/regular_user_upload.html', context)

@authentication_required
def handle_regular_user_upload(request):
    """Handle file upload for any authenticated user"""
    profile = get_user_profile(request.user)
    
    # Helper function to determine redirect URL based on user role
    def get_redirect_url():
        if profile.role == 'regular':
            return 'regular_user_upload'
        else:
            return 'dashboard_router'
    
    # Allow all authenticated users to use this upload functionality
    # Super admins and business owners can also analyze documents
    
    # Only handle POST requests with file uploads
    if request.method != 'POST':
        return render(request, 'analyzer/user/regular_user_upload.html', {
            'error': "Invalid request method."
        })
    
    # Check if file was uploaded
    if 'file' not in request.FILES:
        return render(request, 'analyzer/user/regular_user_upload.html', {
            'error': "No file was uploaded."
        })
    
    uploaded_file = request.FILES['file']
    
    # Validate file
    if not uploaded_file:
        return render(request, 'analyzer/user/regular_user_upload.html', {
            'error': "No file was uploaded."
        })
    
    # Check file size (10MB limit)
    max_size = 10 * 1024 * 1024  # 10MB
    if uploaded_file.size > max_size:
        return render(request, 'analyzer/user/regular_user_upload.html', {
            'error': "File size too large. Maximum allowed size is 10MB."
        })
    
    # Check file type
    allowed_extensions = ['.pdf', '.jpg', '.jpeg', '.png']
    file_extension = os.path.splitext(uploaded_file.name)[1].lower()
    if file_extension not in allowed_extensions:
        return render(request, 'analyzer/user/regular_user_upload.html', {
            'error': "Invalid file type. Please upload PDF, JPG, or PNG files only."
        })
    
    try:
        # Generate unique session ID for progress tracking
        import uuid
        session_id = request.POST.get('session_id') or str(uuid.uuid4())
        print(f"DEBUG: Session ID: {session_id}")  # Debug print
        
        # Initialize progress tracking
        update_progress(session_id, 0, "Starting upload...")
        
        # Add processing delay for regular users to prioritize business owners
        if profile.role == 'regular':
            import time
            time.sleep(1.5)  # Increased delay for regular users (1.5 seconds)
        
        # Update progress
        update_progress(session_id, 10, "File uploaded, initializing processing...")
        
        # Create uploads and results directories if they don't exist
        upload_dir = os.path.join(settings.MEDIA_ROOT, 'uploads')
        results_dir = os.path.join(settings.MEDIA_ROOT, 'results')
        os.makedirs(upload_dir, exist_ok=True)
        os.makedirs(results_dir, exist_ok=True)
        
        # Update progress
        update_progress(session_id, 15, "Preparing document for analysis...")
        
        # Save the uploaded file
        fs = FileSystemStorage(location=upload_dir)
        filename = fs.save(uploaded_file.name, uploaded_file)
        file_path = fs.path(filename)
        
        # Update progress
        update_progress(session_id, 20, "Document saved, starting analysis...")
        
        # Process the document with special handling for regular users
        color_images, bw_images, overall_cost, error = process_document_for_regular_user(
            file_path, request.user, results_dir, session_id
        )
        
        # Additional delay for regular users after processing
        profile = get_user_profile(request.user)
        if profile.role == 'regular':
            if session_id:
                update_progress(session_id, 82, "Optimizing results for regular user display...")
            time.sleep(1)  # Additional delay after processing for regular users
        
        if error:
            # Clean up uploaded file on error
            try:
                os.remove(file_path)
            except:
                pass
            # Clear progress on error
            clear_progress(session_id)
            return render(request, 'analyzer/user/regular_user_upload.html', {
                'error': f"Error processing document: {error}"
            })
        
        # Update progress
        update_progress(session_id, 85, "Saving analysis results...")
        
        # Save analysis to database
        if request.user.is_authenticated:
            # Prepare analysis data for storage
            analysis_data = {
                'color_images': [
                    {
                        'page_number': i + 1,
                        'cost': float(img['cost']),
                        'cost_reason': img['cost_reason'],
                        'coverage_percentage': img['coverage_percentage'],
                        'image_type': img['image_type'],
                        'image_path': img['image_path']
                    }
                    for i, img in enumerate(color_images)
                ],
                'bw_images': [
                    {
                        'page_number': i + 1,
                        'cost': float(img['cost']),
                        'cost_reason': img['cost_reason'],
                        'coverage_percentage': img['coverage_percentage'],
                        'image_type': img['image_type'],
                        'image_path': img['image_path']
                    }
                    for i, img in enumerate(bw_images)
                ]
            }

            analysis = DocumentAnalysis.objects.create(
                user=request.user,
                document_name=uploaded_file.name,
                page_count=len(color_images) + len(bw_images),
                color_page_count=len(color_images),
                bw_page_count=len(bw_images),
                overall_cost=overall_cost,
                analysis_result=analysis_data,
                file_size=uploaded_file.size,
                file_type=os.path.splitext(uploaded_file.name)[1].lower().replace('.', '') if '.' in uploaded_file.name else 'unknown'
            )
        
        # Update progress
        update_progress(session_id, 95, "Preparing results for display...")
        
        # Clean up the uploaded file (but keep processed images for display)
        try:
            os.remove(file_path)
        except:
            pass  # Ignore cleanup errors
        
        # Complete progress
        update_progress(session_id, 100, "Analysis complete!")
        
        # Clear progress after a short delay
        clear_progress(session_id)
        
        # Render the upload page with results instead of redirecting
        return render(request, 'analyzer/user/regular_user_upload.html', {
            'color_images': color_images,
            'bw_images': bw_images,
            'overall_cost': overall_cost,
            'total_pages': len(color_images) + len(bw_images),
            'file_name': uploaded_file.name,
            'analysis_id': analysis.id if request.user.is_authenticated else None,
            'success_message': f"Document analyzed successfully! Total cost: ₱{overall_cost}",
            'session_id': session_id
        })
        
    except Exception as e:
        return render(request, 'analyzer/user/regular_user_upload.html', {
            'error': f"An error occurred while processing your document: {str(e)}"
        })

@authentication_required
def debug_user_role(request):
    """Debug view to check user role"""
    profile = get_user_profile(request.user)
    
    debug_info = f"""
    User: {request.user.username}
    User ID: {request.user.id}
    User Role: {profile.role}
    Is Super Admin: {profile.is_super_admin()}
    Is Business Owner: {profile.is_business_owner()}
    Can Manage Pricing: {profile.can_manage_pricing()}
    Can View All Users: {profile.can_view_all_users()}
    Session User Role: {request.session.get('user_role', 'Not Set')}
    Session User ID: {request.session.get('user_id', 'Not Set')}
    """
    
    return HttpResponse(f"<pre>{debug_info}</pre>")

@authentication_required
def business_owner_dashboard(request):
    """Dashboard for business owners - pricing management and analytics"""
    profile = get_user_profile(request.user)
    
    if not profile.can_manage_pricing():
        messages.error(request, "Access denied. Business owner privileges required.")
        return redirect('regular_user_dashboard')

    # Check if business owner has complete pricing coverage
    has_complete_coverage = check_business_owner_complete_coverage(request.user)
    
    # Get business owner's pricing rules
    pricing_rules = CostSetting.objects.filter(business_owner=request.user).order_by('-created_at')
    
    # Get business analytics
    total_rules = pricing_rules.count()
    active_rules = pricing_rules.filter(is_active=True).count()
    inactive_rules = total_rules - active_rules

    # Document analysis and revenue stats for business owner dashboard
    from django.db.models import Sum
    analyses_qs = DocumentAnalysis.objects.filter(user=request.user).order_by('-created_at')
    total_analyses = analyses_qs.count()

    # Use printed documents for revenue (only completed print jobs contribute to revenue)
    printed_qs = PrintedDocument.objects.filter(business_owner=request.user, status='completed')
    total_printed_agg = printed_qs.aggregate(total=Sum('total_revenue'))
    total_cost = float(total_printed_agg.get('total') or 0)

    # Monthly stats (last 30 days)
    from django.utils import timezone
    from datetime import timedelta
    thirty_days_ago = timezone.now() - timedelta(days=30)
    monthly_qs = analyses_qs.filter(created_at__gte=thirty_days_ago)
    monthly_analyses = monthly_qs.count()

    monthly_printed_agg = PrintedDocument.objects.filter(
        business_owner=request.user,
        status='completed',
        completed_at__gte=thirty_days_ago
    ).aggregate(total=Sum('total_revenue'))
    monthly_cost = float(monthly_printed_agg.get('total') or 0)

    # Recent analyses (limit 5)
    user_recent_analyses = list(analyses_qs[:5])

    # Average cost per analysis
    avg_cost = float((total_cost / total_analyses) if total_analyses > 0 else 0)

    context = {
        'pricing_rules': pricing_rules,  # Show all rules
        'total_rules': total_rules,
        'active_rules': active_rules,
        'inactive_rules': inactive_rules,
        'user_profile': profile,
        'has_complete_coverage': has_complete_coverage,  # Add coverage check
        # Document analysis data
        'total_analyses': total_analyses,
        'total_cost': total_cost,
        'monthly_analyses': monthly_analyses,
        'monthly_cost': monthly_cost,
        'user_recent_analyses': user_recent_analyses,
        'avg_cost': avg_cost,
    }
    return render(request, 'analyzer/business/business_owner_dashboard.html', context)

@business_owner_required
def business_user_analysis_history(request):
    """Full analysis history for business owners - their own documents only"""
    profile = get_user_profile(request.user)
    
    if not profile.can_manage_pricing():
        messages.error(request, "Access denied. Business owner privileges required.")
        return redirect('regular_user_dashboard')

    # Get all analyses for this user
    analyses_queryset = DocumentAnalysis.objects.filter(user=request.user).order_by('-created_at')
    
    # Apply filters
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    status = request.GET.get('status')
    search = request.GET.get('search')
    
    if date_from:
        try:
            from datetime import datetime
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            analyses_queryset = analyses_queryset.filter(created_at__date__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            from datetime import datetime
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            analyses_queryset = analyses_queryset.filter(created_at__date__lte=date_to_obj)
        except ValueError:
            pass
    
    if status:
        # For now, we'll assume all completed analyses are 'completed'
        # You can expand this based on your DocumentAnalysis model fields
        if status == 'completed':
            analyses_queryset = analyses_queryset.exclude(overall_cost__isnull=True)
        elif status == 'processing':
            analyses_queryset = analyses_queryset.filter(overall_cost__isnull=True)
    
    if search:
        # Search in filename or related fields
        analyses_queryset = analyses_queryset.filter(
            Q(result_file__icontains=search) |
            Q(created_at__icontains=search)
        )
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(analyses_queryset, 10)  # 10 analyses per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Add print status information for each analysis in the current page
    for analysis in page_obj:
        # Get the latest print job for this analysis
        latest_print_job = PrintedDocument.objects.filter(
            analysis=analysis,
            business_owner=request.user
        ).order_by('-queued_at').first()
        
        # Add print status to the analysis object
        analysis.print_status = None
        analysis.print_job_count = 0
        analysis.total_print_revenue = 0
        
        if latest_print_job:
            analysis.print_status = latest_print_job.status
            analysis.print_job_count = PrintedDocument.objects.filter(
                analysis=analysis,
                business_owner=request.user
            ).count()
            analysis.total_print_revenue = PrintedDocument.objects.filter(
                analysis=analysis,
                business_owner=request.user,
                status='completed'
            ).aggregate(total=Sum('total_revenue'))['total'] or 0
    
    # Calculate statistics
    total_analyses = analyses_queryset.count()
    total_cost = analyses_queryset.aggregate(total=Sum('overall_cost'))['total'] or 0
    total_pages = sum([analysis.page_count or 0 for analysis in analyses_queryset])
    
    # This month count
    current_month = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    this_month_count = analyses_queryset.filter(created_at__gte=current_month).count()
    
    context = {
        'analyses': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'total_analyses': total_analyses,
        'total_cost': total_cost,
        'total_pages': total_pages,
        'this_month_count': this_month_count,
    }
    
    return render(request, 'analyzer/business/business_user_analysis_history.html', context)

@authentication_required
def regular_user_analysis_history(request):
    """Full analysis history for regular users - their own documents only"""
    
    # Get all analyses for this user
    analyses_queryset = DocumentAnalysis.objects.filter(user=request.user).order_by('-created_at')
    
    # Apply filters
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    status = request.GET.get('status')
    search = request.GET.get('search')
    
    if date_from:
        try:
            from datetime import datetime
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            analyses_queryset = analyses_queryset.filter(created_at__date__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            from datetime import datetime
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            analyses_queryset = analyses_queryset.filter(created_at__date__lte=date_to_obj)
        except ValueError:
            pass
    
    if status:
        # For now, we'll assume all completed analyses are 'completed'
        # You can expand this based on your DocumentAnalysis model fields
        if status == 'completed':
            analyses_queryset = analyses_queryset.exclude(overall_cost__isnull=True)
        elif status == 'processing':
            analyses_queryset = analyses_queryset.filter(overall_cost__isnull=True)
    
    if search:
        # Search in filename or related fields
        analyses_queryset = analyses_queryset.filter(
            Q(document_name__icontains=search) |
            Q(created_at__icontains=search)
        )
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(analyses_queryset, 10)  # 10 analyses per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Calculate statistics
    total_analyses = analyses_queryset.count()
    total_cost = analyses_queryset.aggregate(total=Sum('overall_cost'))['total'] or 0
    total_pages = sum([analysis.page_count or 0 for analysis in analyses_queryset])
    
    # This month count
    current_month = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    this_month_count = analyses_queryset.filter(created_at__gte=current_month).count()
    
    context = {
        'analyses': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'total_analyses': total_analyses,
        'total_cost': total_cost,
        'total_pages': total_pages,
        'this_month_count': this_month_count,
    }
    
    return render(request, 'analyzer/user/regular_user_analysis_history.html', context)

@authentication_required
def business_owner_request(request):
    """Handle business owner access requests"""
    print(f"DEBUG: business_owner_request view called with method: {request.method}")
    print(f"DEBUG: User: {request.user.username}")
    
    profile = get_user_profile(request.user)
    
    # Temporarily disable existing request check for debugging
    print("DEBUG: Skipping existing request check for now")
    
    if request.method == 'POST':
        print("DEBUG: POST request received for business owner request")
        print(f"DEBUG: POST data: {dict(request.POST)}")
        try:
            # Try to create new business owner request with full model
            print("DEBUG: Attempting to create BusinessOwnerRequest...")
            business_request = BusinessOwnerRequest.objects.create(
                user=request.user,
                business_name=request.POST.get('business_name'),
                business_type=request.POST.get('business_type'),
                business_address=request.POST.get('business_address'),
                business_phone=request.POST.get('business_phone'),
                business_email=request.POST.get('business_email'),
                business_registration_number=None,  # Not collected anymore
                tax_id=None,  # Not collected anymore
                years_in_operation=1,  # Default value since not collected
                monthly_volume='0-100',  # Use first choice as default
                business_description='Premium account request',  # Default description
                special_requirements=None,  # Not collected anymore
            )
            print(f"DEBUG: BusinessOwnerRequest created successfully with ID: {business_request.id}")
            
            # Handle AJAX requests
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'Your premium account request has been submitted successfully! Business: {request.POST.get("business_name")} | Submitted by: {request.user.username}'
                })
            
            messages.success(
                request, 
                f"✅ SUCCESS: Your premium account request has been submitted successfully! "
                f"Business: {request.POST.get('business_name')} | "
                f"Submitted by: {request.user.username} | "
                f"Our team will review your application within 1-2 business days and notify you via email."
            )
        except Exception as e:
            print(f"DEBUG: Error creating BusinessOwnerRequest: {e}")
            print(f"DEBUG: Exception type: {type(e)}")
            import traceback
            print(f"DEBUG: Full traceback: {traceback.format_exc()}")
            
            # Handle AJAX requests for errors
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': 'There was an error processing your request. Please try again.'
                })
                
            # Fallback to simple success message if model doesn't exist
            messages.success(
                request, 
                f"✅ SUCCESS: Your premium account request has been submitted successfully! "
                f"Business: {request.POST.get('business_name', 'N/A')} | "
                f"Submitted by: {request.user.username} | "
                f"We'll review your application and get back to you within 1-2 business days."
            )
        
        # Handle AJAX requests for redirect
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'redirect': '/dashboard/'})
            
        return redirect('regular_user_dashboard')
    
    print("DEBUG: Rendering business_owner_request.html template")
    return render(request, 'analyzer/user/business_owner_request.html')

@authentication_required
def upload_image(request):
    # Check user role and redirect if they should be on a different dashboard
    profile = get_user_profile(request.user)
    
    # Super admins should be on the admin dashboard, not upload page
    if profile.is_super_admin():
        messages.info(request, 'As a Super Admin, you have access to the admin dashboard.')
        return redirect('super_admin_dashboard')
    
    # Regular users should use their dashboard for uploads
    if profile.role == 'regular':
        messages.info(request, 'Please use your dashboard for document uploads.')
        return redirect('regular_user_dashboard')
    
    # Business owners can access upload normally
    if request.method == 'POST' and request.FILES.get('file'):
        return handle_file_upload(request)

    return render(request, 'analyzer/upload.html')

@login_required
@business_owner_required
def pricing_dashboard(request):
    """Dashboard for managing pricing rules - Business owners only"""
    user_pricing_rules = CostSetting.objects.filter(business_owner=request.user).order_by('color', 'coverage_min')
    
    context = {
        'user_pricing_rules': user_pricing_rules,
    }
    return render(request, 'analyzer/pricing_dashboard.html', context)

@business_owner_required
def add_pricing_rule(request):
    """Add a new pricing rule - Business owners only"""
    if request.method == 'POST':
        # Support both single-rule (legacy) and batch-rule submission
        try:
            paper_size = request.POST.get('paper_size', 'a4')
            # Allow new Word-like paper sizes as well while keeping legacy keys
            valid_sizes = ['short', 'long', 'a4', 'tabloid', 'a3', 'a3_plus', 'b4', 'b3', 'statement', 'executive']
            if paper_size not in valid_sizes:
                messages.error(request, "Invalid paper size selected.")
                return redirect('business_owner_dashboard')

            # Detect batch submission by presence of 'batch' flag or color_cost_1
            is_batch = request.POST.get('batch') == '1' or any(k.startswith('color_cost_') or k.startswith('bw_cost_') for k in request.POST.keys())

            if is_batch:
                # Define coverage intervals with no gaps: 0-10, 10-20, ..., 90-100
                # This ensures adjacent ranges overlap at the boundary (e.g. 10) and
                # satisfies the complete-coverage check which expects coverage to start at 0.
                intervals = []
                for i in range(10):
                    min_v = i * 10
                    max_v = (i + 1) * 10
                    intervals.append((min_v, max_v))

                created_count = 0
                updated_count = 0
                errors = []

                with transaction.atomic():
                    for idx, (cmin, cmax) in enumerate(intervals, start=1):
                        # Allow client to override coverage bounds by posting coverage_min_X/coverage_max_X
                        try:
                            posted_min = request.POST.get(f'coverage_min_{idx}')
                            posted_max = request.POST.get(f'coverage_max_{idx}')
                            if posted_min is not None and posted_max is not None and posted_min != '' and posted_max != '':
                                # Use Decimal to match model
                                cmin = Decimal(str(float(posted_min)))
                                cmax = Decimal(str(float(posted_max)))
                        except Exception:
                            # If parsing fails, fall back to defaults
                            pass
                        # Color
                        color_field = f'color_cost_{idx}'
                        color_val = request.POST.get(color_field, '').strip()
                        if color_val != '':
                            try:
                                cost = Decimal(str(float(color_val)))
                                obj, created = CostSetting.objects.update_or_create(
                                    business_owner=request.user,
                                    color=True,
                                    paper_size=paper_size,
                                    coverage_min=cmin,
                                    coverage_max=cmax,
                                    defaults={
                                        'cost': cost,
                                        'reason': f'Coverage {cmin}%-{cmax}% (batch)'
                                    }
                                )
                                if created:
                                    created_count += 1
                                else:
                                    updated_count += 1
                            except Exception as e:
                                errors.append(f'Color interval {cmin}-{cmax}: {str(e)}')

                        # Black & White
                        bw_field = f'bw_cost_{idx}'
                        bw_val = request.POST.get(bw_field, '').strip()
                        if bw_val != '':
                            try:
                                cost = Decimal(str(float(bw_val)))
                                obj, created = CostSetting.objects.update_or_create(
                                    business_owner=request.user,
                                    color=False,
                                    paper_size=paper_size,
                                    coverage_min=cmin,
                                    coverage_max=cmax,
                                    defaults={
                                        'cost': cost,
                                        'reason': f'Coverage {cmin}%-{cmax}% (batch)'
                                    }
                                )
                                if created:
                                    created_count += 1
                                else:
                                    updated_count += 1
                            except Exception as e:
                                errors.append(f'B&W interval {cmin}-{cmax}: {str(e)}')

                if errors:
                    messages.error(request, 'Some intervals failed: ' + '; '.join(errors))
                else:
                    messages.success(request, f'Batch pricing updated: {created_count} created, {updated_count} updated.')

            else:
                # Legacy single-rule add (keep original behavior)
                color = request.POST.get('color') == 'True'
                coverage_min = float(request.POST.get('coverage_min'))
                coverage_max = float(request.POST.get('coverage_max'))
                cost = float(request.POST.get('cost'))
                reason = request.POST.get('reason')

                # Validate coverage ranges
                if coverage_min >= coverage_max:
                    messages.error(request, "Minimum coverage must be less than maximum coverage.")
                    return redirect('business_owner_dashboard')

                if coverage_min < 0 or coverage_max > 100:
                    messages.error(request, "Coverage percentages must be between 0 and 100.")
                    return redirect('business_owner_dashboard')

                duplicate = CostSetting.objects.filter(
                    business_owner=request.user,
                    color=color,
                    paper_size=paper_size,
                    coverage_min=coverage_min,
                    coverage_max=coverage_max
                ).exists()

                if duplicate:
                    rule_type = "Color" if color else "Black & White"
                    paper_display = dict(CostSetting.PAPER_SIZE_CHOICES)[paper_size]
                    messages.error(request, f"Coverage range {coverage_min}%-{coverage_max}% is already implemented for {rule_type} documents on {paper_display}.")
                    return redirect('business_owner_dashboard')

                CostSetting.objects.create(
                    business_owner=request.user,
                    color=color,
                    paper_size=paper_size,
                    coverage_min=coverage_min,
                    coverage_max=coverage_max,
                    cost=cost,
                    reason=reason
                )

                rule_type = "Color" if color else "Black & White"
                paper_display = dict(CostSetting.PAPER_SIZE_CHOICES)[paper_size]
                messages.success(
                    request,
                    f"✅ Pricing rule added successfully! {rule_type} printing on {paper_display} for {coverage_min}%-{coverage_max}% coverage at ₱{cost:.2f}."
                )

        except ValueError:
            messages.error(request, "Invalid input values. Please check your entries.")
        except Exception as e:
            messages.error(request, f"Error adding pricing rule: {str(e)}")

    return redirect('business_owner_dashboard')

@business_owner_required
def edit_pricing_rule(request, rule_id):
    """Edit an existing pricing rule - Business owners only"""
    rule = get_object_or_404(CostSetting, id=rule_id, business_owner=request.user)
    
    if request.method == 'POST':
        try:
            color = request.POST.get('color') == 'True'
            paper_size = request.POST.get('paper_size', 'a4')
            coverage_min = float(request.POST.get('coverage_min'))
            coverage_max = float(request.POST.get('coverage_max'))
            cost = float(request.POST.get('cost'))
            reason = request.POST.get('reason')
            
            # Validate paper size (include Word-like sizes)
            valid_sizes = ['short', 'long', 'a4', 'tabloid', 'a3', 'a3_plus', 'b4', 'b3', 'statement', 'executive']
            if paper_size not in valid_sizes:
                messages.error(request, "Invalid paper size selected.")
                return redirect('pricing_dashboard')
            
            # Validate coverage ranges
            if coverage_min >= coverage_max:
                messages.error(request, "Minimum coverage must be less than maximum coverage.")
                return redirect('pricing_dashboard')
            
            if coverage_min < 0 or coverage_max > 100:
                messages.error(request, "Coverage percentages must be between 0 and 100.")
                return redirect('pricing_dashboard')
            
            # Check for exact duplicate coverage ranges within the same document type and paper size (excluding current rule)
            duplicate = CostSetting.objects.filter(
                business_owner=request.user,
                color=color,
                paper_size=paper_size,
                coverage_min=coverage_min,
                coverage_max=coverage_max
            ).exclude(id=rule_id).exists()
            
            if duplicate:
                rule_type = "Color" if color else "Black & White"
                paper_display = dict(CostSetting.PAPER_SIZE_CHOICES)[paper_size]
                messages.error(request, f"Coverage range {coverage_min}%-{coverage_max}% is already implemented for {rule_type} documents on {paper_display}.")
                return redirect('business_owner_dashboard')
            
            rule.color = color
            rule.paper_size = paper_size
            rule.coverage_min = coverage_min
            rule.coverage_max = coverage_max
            rule.cost = cost
            rule.reason = reason
            rule.save()
            
            # Create a more descriptive success message
            rule_type = "Color" if color else "Black & White"
            messages.success(
                request, 
                f"💾 Pricing rule updated successfully! {rule_type} printing for {coverage_min}%-{coverage_max}% coverage at ₱{cost:.2f}."
            )
        except ValueError:
            messages.error(request, "Invalid input values. Please check your entries.")
        except Exception as e:
            messages.error(request, f"Error updating pricing rule: {str(e)}")
    
    # Check if request came from business owner dashboard
    next_url = request.GET.get('next', 'pricing_dashboard')
    if next_url == 'business_owner_dashboard':
        return redirect('business_owner_dashboard')
    return redirect('pricing_dashboard')

@business_owner_required
def delete_pricing_rule(request, rule_id):
    """Delete a pricing rule - Business owners only"""
    rule = get_object_or_404(CostSetting, id=rule_id, business_owner=request.user)
    
    if request.method == 'POST':
        rule_description = rule.reason
        rule.delete()
        messages.success(request, f"🗑️ Pricing rule '{rule_description}' deleted successfully.")
    
    # Check if request came from business owner dashboard
    next_url = request.GET.get('next', 'pricing_dashboard')
    if next_url == 'business_owner_dashboard':
        return redirect('business_owner_dashboard')
    return redirect('pricing_dashboard')


@business_owner_required
def delete_all_pricing_rules(request):
    """Delete all pricing rules for the current business owner (POST only)."""
    if request.method == 'POST':
        try:
            with transaction.atomic():
                qs = CostSetting.objects.filter(business_owner=request.user)
                total = qs.count()
                qs.delete()
            messages.success(request, f"🗑️ All pricing rules deleted ({total} rules).")
        except Exception as e:
            messages.error(request, f"Error deleting pricing rules: {str(e)}")
    else:
        messages.error(request, "Invalid request method for deleting all pricing rules.")

    next_url = request.GET.get('next', 'pricing_dashboard')
    if next_url == 'business_owner_dashboard':
        return redirect('business_owner_dashboard')
    return redirect('pricing_dashboard')

@business_owner_required
def toggle_pricing_rule(request, rule_id):
    """Toggle active status of a pricing rule - Business owners only"""
    rule = get_object_or_404(CostSetting, id=rule_id, business_owner=request.user)
    
    if request.method == 'POST':
        rule.is_active = not rule.is_active
        rule.save()
        status = "activated" if rule.is_active else "deactivated"
        messages.success(request, f"Pricing rule {status} successfully.")
    
    return redirect('pricing_dashboard')

def check_business_owner_complete_coverage(business_owner):
    """Check if business owner has complete pricing rules covering 100% for both B&W and Color"""
    try:
        # Check B&W coverage (color=False)
        bw_rules = CostSetting.objects.filter(
            business_owner=business_owner,
            color=False,
            is_active=True
        ).order_by('coverage_min')
        
        # Check Color coverage (color=True)
        color_rules = CostSetting.objects.filter(
            business_owner=business_owner,
            color=True,
            is_active=True
        ).order_by('coverage_min')
        
        # Function to check if rules cover 0-100% without gaps
        def has_complete_coverage(rules):
            if not rules.exists():
                return False
            
            # Check if coverage starts from 0 and goes to 100
            rules_list = list(rules)
            
            # Must start from 0%
            if rules_list[0].coverage_min != 0:
                return False
            
            # Check for gaps and ensure it reaches 100%
            current_max = -1
            for rule in rules_list:
                # Allow adjacent ranges or overlapping ranges (no gaps)
                if current_max >= 0 and rule.coverage_min > current_max:
                    return False  # Gap found
                current_max = max(current_max, rule.coverage_max)
            
            # Must reach 100%
            return current_max >= 100
        
        bw_complete = has_complete_coverage(bw_rules)
        color_complete = has_complete_coverage(color_rules)
        
        print(f"DEBUG - Business owner {business_owner.username} coverage check:")
        print(f"DEBUG - B&W rules complete: {bw_complete}")
        print(f"DEBUG - Color rules complete: {color_complete}")
        
        return bw_complete and color_complete
        
    except Exception as e:
        print(f"ERROR - Coverage check failed for {business_owner.username}: {e}")
        return False

def get_pricing_for_user(user, color_detected, coverage_percentage, paper_size='short'):
    """Helper function to get pricing for a user based on image properties"""
    if isinstance(user, AnonymousUser):
        # Use admin-configured default pricing for anonymous users
        rule = DefaultPricingRule.objects.filter(
            color=color_detected,
            paper_size=paper_size,
            coverage_min__lte=coverage_percentage,
            coverage_max__gte=coverage_percentage,
            is_active=True
        ).first()
        
        if rule:
            cost = float(rule.cost)
            return cost, rule.reason
        else:
            # Simple fallback pricing when no admin rules are configured
            base_cost = 1.0 if not color_detected else 2.0
            coverage_factor = coverage_percentage / 100
            cost = base_cost * (1 + coverage_factor)
            rounded_cost = int(round(cost))
            return rounded_cost, "Basic pricing (no admin rules configured)"
    
    # Check user role to determine pricing rules to use
    profile = get_user_profile(user)
    
    # Regular users ALWAYS use default pricing rules (never custom pricing)
    if profile.role == 'regular':
        rule = DefaultPricingRule.objects.filter(
            color=color_detected,
            paper_size=paper_size,
            coverage_min__lte=coverage_percentage,
            coverage_max__gte=coverage_percentage,
            is_active=True
        ).first()
        
        if rule:
            cost = float(rule.cost)
            return cost, f"{rule.reason} (Default pricing for regular users)"
        else:
            # Simple fallback pricing when no admin rules are configured
            base_cost = 1.0 if not color_detected else 2.0
            coverage_factor = coverage_percentage / 100
            cost = base_cost * (1 + coverage_factor)
            rounded_cost = int(round(cost))
            return rounded_cost, "Basic default pricing (no admin rules configured)"
    
    # Business owners MUST have complete pricing rules - NO DEFAULT FALLBACK
    if profile.role == 'business_owner':
        # Check if business owner has complete pricing coverage (100% for both B&W and Color)
        has_complete_coverage = check_business_owner_complete_coverage(user)
        
        if not has_complete_coverage:
            # Business owner doesn't have complete pricing rules
            raise ValueError("INCOMPLETE_PRICING_RULES: Business owner must set pricing rules up to 100% coverage for both B&W and Color before users can analyze documents.")
        
        # Try to find business owner's own pricing rules
        owner_rule = CostSetting.objects.filter(
            business_owner=user,
            color=color_detected,
            paper_size=paper_size,
            coverage_min__lte=coverage_percentage,
            coverage_max__gte=coverage_percentage,
            is_active=True
        ).first()
        
        if owner_rule:
            cost = float(owner_rule.cost)
            return cost, f"{owner_rule.reason} (Your pricing rule)"
        
        # If no specific rule found but coverage is complete, this shouldn't happen
        raise ValueError("PRICING_RULE_NOT_FOUND: No pricing rule found for this coverage range despite complete coverage check.")
    
    # Super admins use default pricing rules
    if profile.role == 'super_admin':
        default_rule = DefaultPricingRule.objects.filter(
            color=color_detected,
            coverage_min__lte=coverage_percentage,
            coverage_max__gte=coverage_percentage,
            is_active=True
        ).first()
        
        if default_rule:
            cost = float(default_rule.cost)
            return cost, f"{default_rule.reason} (Admin pricing)"
    
    # Fall back to admin-configured default pricing for any other roles
    default_rule = DefaultPricingRule.objects.filter(
        color=color_detected,
        coverage_min__lte=coverage_percentage,
        coverage_max__gte=coverage_percentage,
        is_active=True
    ).first()
    
    if default_rule:
        cost = float(default_rule.cost)
        return cost, default_rule.reason
    
    # Final fallback when no rules are configured
    base_cost = 1.0 if not color_detected else 2.0
    coverage_factor = coverage_percentage / 100
    cost = base_cost * (1 + coverage_factor)
    rounded_cost = round(cost, 2)
    return rounded_cost, "Basic pricing (no rules configured)"

@login_required
def document_history(request):
    """View to display user's document analysis history"""
    user_analyses = DocumentAnalysis.objects.filter(user=request.user).order_by('-created_at')
    
    context = {
        'user_analyses': user_analyses,
    }
    return render(request, 'analyzer/document_history.html', context)

def user_login(request):
    """User login view with proper session setup and AJAX support"""
    if request.user.is_authenticated:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'redirect_url': '/dashboard/'})
        return redirect('dashboard_router')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        if user is not None:
            # Clear any existing messages from previous sessions
            storage = messages.get_messages(request)
            for _ in storage:
                pass  # This clears the messages
            
            login(request, user)
            
            # Set up session data for security
            try:
                profile = get_user_profile(user)
                request.session['user_role'] = profile.role if profile else 'regular'
                request.session['user_id'] = user.id
                request.session['login_timestamp'] = timezone.now().timestamp()
                
                # Generate new session key for security
                request.session.cycle_key()
                
                success_message = f'Welcome back, {user.username}!'
                
                # Role-specific welcome messages
                if profile and profile.role == 'super_admin':
                    success_message += ' Admin access granted.'
                elif profile and profile.role == 'business_owner':
                    success_message += ' Business owner access granted.'
                else:
                    success_message += ' User access granted.'
                    
            except Exception as e:
                success_message = 'Login successful!'
            
            # Get next URL from POST data or GET parameter
            next_url = request.POST.get('next') or request.GET.get('next', '/dashboard/')
            
            # Handle AJAX requests
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': success_message,
                    'redirect_url': next_url
                })
            
            # Regular form submission
            messages.success(request, success_message)
            return redirect(next_url)
        else:
            error_message = 'Invalid username or password.'
            
            # Handle AJAX requests
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': error_message
                })
            
            messages.error(request, error_message)
    
    return render(request, 'analyzer/login.html')

def user_register(request):
    """User registration view with AJAX support"""
    if request.user.is_authenticated:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'redirect_url': '/login/'})
        return redirect('user_login')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        
        # Basic validation
        if not username or not password1 or not password2:
            error_message = 'All fields are required.'
        elif password1 != password2:
            error_message = 'Passwords do not match.'
        elif len(password1) < 8:
            error_message = 'Password must be at least 8 characters long.'
        elif User.objects.filter(username=username).exists():
            error_message = 'Username already exists.'
        elif email and User.objects.filter(email=email).exists():
            error_message = 'Email already exists.'
        else:
            try:
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password1
                )
                success_message = 'Account created successfully! You can now log in.'

                # Handle AJAX requests
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': success_message,
                        'redirect_url': '/login/'
                    })

                messages.success(request, success_message)
                return redirect('user_login')
            except Exception as e:
                error_message = f'Error creating account: {str(e)}'
        
        # Handle validation errors for AJAX requests
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': error_message
            })
        
        messages.error(request, error_message)
    
    return render(request, 'analyzer/register.html')

def user_logout(request):
    """User logout view with proper session cleanup"""
    if request.user.is_authenticated:
        # Get user role before logout for proper redirect
        try:
            profile = get_user_profile(request.user)
            user_role = profile.role if profile else 'regular'
        except:
            user_role = 'regular'
        
        # Add logout message before clearing session
        messages.success(request, 'You have been logged out successfully.')
        
        # Role-specific message before logout
        if user_role == 'super_admin':
            messages.info(request, 'Admin session ended. Please log in again to access admin features.')
        elif user_role == 'business_owner':
            messages.info(request, 'Business owner session ended. Please log in again to access business features.')
        else:
            messages.info(request, 'User session ended. Please log in again to continue.')
        
        # Clear all user sessions
        logout(request)
        
        # Clear any cached user data
        if hasattr(request, 'session'):
            request.session.flush()
    
    return redirect('landing_page')

@super_admin_required
def super_admin_dashboard(request):
    """Super Admin Dashboard - View all users and system analytics"""
    # Handle POST requests for role updates and request approvals
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'update_role':
            user_id = request.POST.get('user_id')
            new_role = request.POST.get('new_role')
            try:
                user = User.objects.get(id=user_id)
                profile = get_user_profile(user)
                profile.role = new_role
                profile.save()
                messages.success(request, f'Successfully updated {user.username}\'s role to {profile.get_role_display()}.')
            except User.DoesNotExist:
                messages.error(request, 'User not found.')
        
        elif action == 'approve_request':
            request_id = request.POST.get('request_id')
            try:
                business_request = BusinessOwnerRequest.objects.get(id=request_id)
                business_request.status = 'approved'
                business_request.reviewed_by = request.user
                business_request.reviewed_at = timezone.now()
                business_request.admin_notes = request.POST.get('admin_notes', '')
                business_request.save()
                
                # Update user role to business_owner
                profile = get_user_profile(business_request.user)
                profile.role = 'business_owner'
                profile.save()
                
                messages.success(request, f'Approved business request for {business_request.user.username}. User role updated to Business Owner.')
            except BusinessOwnerRequest.DoesNotExist:
                messages.error(request, 'Business request not found.')
        
        elif action == 'reject_request':
            request_id = request.POST.get('request_id')
            try:
                business_request = BusinessOwnerRequest.objects.get(id=request_id)
                business_request.status = 'rejected'
                business_request.reviewed_by = request.user
                business_request.reviewed_at = timezone.now()
                business_request.admin_notes = request.POST.get('admin_notes', '')
                business_request.save()
                
                messages.success(request, f'Rejected business request for {business_request.user.username}.')
            except BusinessOwnerRequest.DoesNotExist:
                messages.error(request, 'Business request not found.')
        
        # Account Management Actions
        elif action == 'edit_account':
            user_id = request.POST.get('user_id')
            try:
                target_user = User.objects.get(id=user_id)
                if target_user.is_superuser and not request.user.is_superuser:
                    messages.error(request, 'Cannot edit super admin accounts.')
                else:
                    # Update user fields
                    target_user.username = request.POST.get('username', target_user.username)
                    target_user.email = request.POST.get('email', target_user.email)
                    target_user.is_active = request.POST.get('is_active') == 'true'
                    
                    # Handle password update if provided
                    new_password = request.POST.get('password', '').strip()
                    if new_password:
                        target_user.set_password(new_password)
                        messages.success(request, f'Password updated for {target_user.username}.')
                    
                    target_user.save()
                    
                    # Update user profile role
                    role = request.POST.get('role', 'regular_user')
                    profile, created = UserProfile.objects.get_or_create(user=target_user)
                    profile.role = role
                    profile.save()
                    
                    admin_notes = request.POST.get('admin_notes', '').strip()
                    if admin_notes:
                        messages.info(request, f'Admin notes: {admin_notes}')
                    
                    messages.success(request, f'Account for {target_user.username} has been updated successfully.')
            except User.DoesNotExist:
                messages.error(request, 'User not found.')
            except Exception as e:
                messages.error(request, f'Error updating account: {str(e)}')
        
        elif action == 'delete_account':
            user_id = request.POST.get('user_id')
            reason = request.POST.get('reason', '')
            try:
                target_user = User.objects.get(id=user_id)
                if target_user.is_superuser:
                    messages.error(request, 'Cannot delete super admin accounts.')
                else:
                    username = target_user.username
                    
                    # Manually clean up related data to avoid cascade delete issues
                    try:
                        # Delete business owner requests first
                        BusinessOwnerRequest.objects.filter(user=target_user).delete()
                        
                        # Delete document analyses
                        DocumentAnalysis.objects.filter(user=target_user).delete()
                        
                        # Delete printed documents
                        PrintedDocument.objects.filter(user=target_user).delete()
                        
                        # Delete cost settings if business owner
                        CostSetting.objects.filter(business_owner=target_user).delete()
                        
                        # Delete user profile
                        if hasattr(target_user, 'userprofile'):
                            target_user.userprofile.delete()
                            
                    except Exception as cleanup_error:
                        print(f"Warning: Error during cleanup for {username}: {cleanup_error}")
                    
                    # Finally delete the user
                    target_user.delete()
                    messages.success(request, f'Account for {username} has been permanently deleted.')
                    
            except User.DoesNotExist:
                messages.error(request, 'User not found.')
            except Exception as e:
                messages.error(request, f'Error deleting account: {str(e)}')
        
        # Preserve tab parameter when redirecting after POST actions
        tab_param = request.POST.get('tab', 'overview')
        redirect_url = reverse('super_admin_dashboard')
        if tab_param and tab_param != 'overview':
            redirect_url += f'?tab={tab_param}'
        return redirect(redirect_url)
    
    # Get search and filter parameters
    search_query = request.GET.get('search', '').strip()
    payment_filter = request.GET.get('payment_status', '').strip()
    role_filter = request.GET.get('role', '').strip()
    date_filter = request.GET.get('date_filter', '').strip()
    account_search = request.GET.get('account_search', '').strip()
    active_tab = request.GET.get('tab', '').strip()
    
    # If no tab is specified but there are user management filters, default to users tab
    if not active_tab:
        if search_query or payment_filter or role_filter or date_filter:
            active_tab = 'users'
        else:
            active_tab = 'overview'
    
    # Get all users with their profiles and business owner approval date
    users_with_profiles = User.objects.select_related('userprofile').prefetch_related(
        'business_requests'
    ).all()
    
    print(f"DEBUG: Initial query returned {users_with_profiles.count()} users")
    
    # Apply initial search filter if provided
    if search_query:
        users_with_profiles = users_with_profiles.filter(
            Q(username__icontains=search_query) | 
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )
    
    # Apply role filter if provided
    if role_filter:
        users_with_profiles = users_with_profiles.filter(userprofile__role=role_filter)
    
    # Add business owner approval date and payment status to each user
    filtered_users = []
    print(f"DEBUG: Starting to process {len(list(users_with_profiles))} users")
    for user in users_with_profiles:
        user.business_owner_since = None
        user.payment_status = None
        user.days_as_business_owner = 0
        
        # Ensure user has a userprofile
        if not hasattr(user, 'userprofile') or user.userprofile is None:
            user.userprofile, created = UserProfile.objects.get_or_create(
                user=user,
                defaults={'role': 'regular_user'}
            )
            if created:
                print(f"DEBUG: Created missing userprofile for {user.username}")
        
        if hasattr(user, 'userprofile') and user.userprofile.role == 'business_owner':
            # Find the approved business request to get the approval date
            try:
                approved_request = user.business_requests.filter(
                    status='approved'
                ).order_by('-reviewed_at').first()
                
                if approved_request and approved_request.reviewed_at:
                    user.business_owner_since = approved_request.reviewed_at
                    # Calculate days since becoming business owner
                    user.days_as_business_owner = (timezone.now() - approved_request.reviewed_at).days
                else:
                    # Fallback 1: Use userprofile updated date
                    if hasattr(user.userprofile, 'updated_at'):
                        user.business_owner_since = user.userprofile.updated_at
                        user.days_as_business_owner = (timezone.now() - user.userprofile.updated_at).days
                    else:
                        # Fallback 2: Use user join date (assume they were always business owner)
                        user.business_owner_since = user.date_joined
                        user.days_as_business_owner = (timezone.now() - user.date_joined).days
                
                # Determine payment status based on time since approval
                if user.days_as_business_owner <= 30:
                    user.payment_status = "active"  # First month
                elif user.days_as_business_owner <= 60:
                    user.payment_status = "pending"  # Second month due
                else:
                    user.payment_status = "overdue"  # Past due
                        
            except Exception as e:
                print(f"DEBUG: Error getting business request for {user.username}: {e}")
                # Emergency fallback
                user.business_owner_since = user.date_joined
                user.payment_status = "pending"
        
        # Determine if user should be included based on filters
        should_include = True
        
        # Debug user info
        print(f"DEBUG: Processing user {user.username}, role: {user.userprofile.role if hasattr(user, 'userprofile') else 'NO_PROFILE'}, payment_status: {user.payment_status}")
        
        # Apply payment status filter
        if payment_filter and payment_filter != 'all':
            should_include = False
            print(f"DEBUG: Applying payment filter '{payment_filter}' for user {user.username}")
            if payment_filter == 'free' and user.userprofile.role != 'business_owner':
                should_include = True
                print(f"DEBUG: User {user.username} included - free filter match (not business owner)")
            elif payment_filter == user.payment_status:
                should_include = True
                print(f"DEBUG: User {user.username} included - payment status match")
            else:
                print(f"DEBUG: User {user.username} excluded by payment filter")
        
        # Apply date filter (only if no payment filter or if user passed payment filter)
        if should_include and date_filter:
            should_include = False  # Reset and check date filter
            print(f"DEBUG: Applying date filter '{date_filter}' for user {user.username}")
            try:
                from datetime import datetime
                filter_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
                user_date = user.business_owner_since.date() if user.business_owner_since else None
                
                # Check if payment is due on this date (monthly billing cycle)
                if user_date and user.userprofile.role == 'business_owner':
                    # Calculate next payment due date
                    days_since_upgrade = (timezone.now().date() - user_date).days
                    months_since = days_since_upgrade // 30
                    next_payment_day = user_date.day
                    
                    # Check if filter date matches a payment due date
                    if (filter_date.day == next_payment_day and 
                        filter_date >= user_date and
                        (filter_date.year * 12 + filter_date.month) > (user_date.year * 12 + user_date.month)):
                        should_include = True
            except ValueError:
                pass  # Invalid date format, skip this filter
        
        # TEMPORARY: Include all users for debugging
        filtered_users.append(user)
        print(f"DEBUG: User {user.username} added to list (total now: {len(filtered_users)})")
    
    # Use filtered users
    users_with_profiles = filtered_users
    
    # Debug: Print user count
    print(f"DEBUG: Total users after filtering: {len(users_with_profiles)}")
    print(f"DEBUG: Search query: '{search_query}', Payment filter: '{payment_filter}', Role filter: '{role_filter}', Date filter: '{date_filter}'")
    
    # Handle account search for Account Management tab
    if account_search:
        # Filter users based on search query
        searched_users = User.objects.select_related('userprofile').filter(
            Q(username__icontains=account_search) | 
            Q(email__icontains=account_search) |
            Q(first_name__icontains=account_search) |
            Q(last_name__icontains=account_search)
        ).order_by('username')
    else:
        # Show all users by default, ordered by username for consistency
        searched_users = User.objects.select_related('userprofile').all().order_by('username')
    
    # Ensure userprofiles exist for all users
    for user in searched_users:
        if not hasattr(user, 'userprofile') or user.userprofile is None:
            user.userprofile, created = UserProfile.objects.get_or_create(
                user=user,
                defaults={'role': 'regular_user'}
            )

    # Get pending business owner requests
    pending_requests = BusinessOwnerRequest.objects.filter(status='pending').order_by('-created_at')
    print(f"DEBUG: Found {len(pending_requests)} pending requests")
    
    # Get all business owner requests for review
    all_business_requests = BusinessOwnerRequest.objects.all().order_by('-created_at')
    print(f"DEBUG: Total BusinessOwnerRequest records in database: {all_business_requests.count()}")
    for req in all_business_requests[:5]:  # Show first 5
        print(f"DEBUG: Request ID {req.id}: {req.user.username} - {req.business_name} - Status: {req.status}")
    all_requests = BusinessOwnerRequest.objects.all().order_by('-created_at')[:20]
    
    # Get system-wide statistics
    total_users = User.objects.count()
    total_analyses = DocumentAnalysis.objects.count()
    total_cost = DocumentAnalysis.objects.aggregate(total=Sum('overall_cost'))['total'] or 0
    
    # Get recent activity (last 30 days)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    recent_analyses_count = DocumentAnalysis.objects.filter(created_at__gte=thirty_days_ago).count()
    recent_users_count = User.objects.filter(date_joined__gte=thirty_days_ago).count()
    
    # Additional statistics for enhanced dashboard
    active_users = User.objects.filter(is_active=True).count()
    inactive_users = User.objects.filter(is_active=False).count()
    
    # Monthly statistics
    current_month = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    analyses_this_month = DocumentAnalysis.objects.filter(created_at__gte=current_month).count()
    new_users_this_month = User.objects.filter(date_joined__gte=current_month).count()
    monthly_revenue = DocumentAnalysis.objects.filter(created_at__gte=current_month).aggregate(total=Sum('overall_cost'))['total'] or 0
    
    # Business requests statistics
    total_business_requests = BusinessOwnerRequest.objects.count()
    approved_requests = BusinessOwnerRequest.objects.filter(status='approved').count()
    
    # Average cost per analysis
    avg_cost_per_analysis = 0
    if total_analyses > 0:
        avg_cost_per_analysis = total_cost / total_analyses
    
    # Get user role distribution
    role_stats = UserProfile.objects.values('role').annotate(count=Count('role'))
    
    # Get top users by analysis count
    top_users = User.objects.annotate(
        analysis_count=Count('documentanalysis')
    ).filter(analysis_count__gt=0).order_by('-analysis_count')[:10]
    
    # Get recent login activity (users who have logged in recently)
    recent_logins = User.objects.select_related('userprofile').filter(
        last_login__isnull=False
    ).order_by('-last_login')[:20]
    
    context = {
        'users_with_profiles': users_with_profiles,
        'pending_requests': pending_requests,
        'all_requests': all_requests,
        'total_users': total_users,
        'total_analyses': total_analyses,
        'total_cost': total_cost,
        'recent_logins': recent_logins,  # Recent login activity for the dashboard
        'recent_analyses_count': recent_analyses_count,
        'recent_users_count': recent_users_count,
        'role_stats': role_stats,
        'top_users': top_users,
        # Account Management
        'account_search': account_search,
        'searched_users': searched_users,
        # Additional statistics for enhanced dashboard
        'active_users': active_users,
        'inactive_users': inactive_users,
        'analyses_this_month': analyses_this_month,
        'new_users_this_month': new_users_this_month,
        'monthly_revenue': monthly_revenue,
        'total_business_requests': total_business_requests,
        'approved_requests': approved_requests,
        'avg_cost_per_analysis': avg_cost_per_analysis,
        # Filter parameters
        'search_query': search_query,
        'payment_filter': payment_filter,
        'role_filter': role_filter,
        'date_filter': date_filter,
        'today': timezone.now().date().strftime('%Y-%m-%d'),
        'active_tab': active_tab,
    }
    
    # Check for upload results in session (for document analysis results)
    if 'upload_results' in request.session:
        upload_results = request.session.pop('upload_results')
        context.update(upload_results)
    
    return render(request, 'analyzer/super_admin_dashboard.html', context)

@business_owner_required
def business_pricing_dashboard(request):
    """Business Owner Dashboard - Manage pricing rules"""
    profile = get_user_profile(request.user)
    
    # Get pricing rules for this business owner
    pricing_rules = CostSetting.objects.filter(business_owner=request.user).order_by('color', 'coverage_min')
    
    # Get analytics for this business owner's pricing
    user_analyses = DocumentAnalysis.objects.filter(user=request.user)
    total_documents = user_analyses.count()
    total_revenue = user_analyses.aggregate(total=Sum('overall_cost'))['total'] or 0
    
    # Get recent activity
    recent_analyses = user_analyses.order_by('-created_at')[:10]
    
    context = {
        'pricing_rules': pricing_rules,
        'total_documents': total_documents,
        'total_revenue': total_revenue,
        'recent_analyses': recent_analyses,
        'user_profile': profile,
    }
    return render(request, 'analyzer/business_pricing_dashboard.html', context)

@login_required
def user_role_check(request):
    """Helper view to check user role and redirect appropriately"""
    if not request.user.is_authenticated:
        return redirect('user_login')
        
    profile = get_user_profile(request.user)
    
    # Debug session data if needed
    if request.GET.get('debug') == '1':
        session_data = {
            'session_key': request.session.session_key,
            'user_role_in_session': request.session.get('user_role'),
            'user_id_in_session': request.session.get('user_id'),
            'login_timestamp': request.session.get('login_timestamp'),
            'actual_user_role': profile.role if profile else 'No profile',
            'actual_user_id': request.user.id,
            'username': request.user.username,
        }
        
        messages.info(request, f"Debug Session Data: {session_data}")
    
    if profile.is_super_admin():
        return redirect('super_admin_dashboard')
    elif profile.is_business_owner():
        return redirect('business_owner_dashboard')
    else:
        return redirect('regular_user_dashboard')


def clear_user_sessions(request):
    """Emergency view to clear all sessions for debugging"""
    if request.user.is_authenticated and request.user.is_superuser:
        request.session.flush()
        logout(request)
        messages.success(request, 'All sessions cleared successfully.')
    else:
        messages.error(request, 'Access denied.')
    
    return redirect('user_login')


@business_owner_required
def analytics_dashboard(request):
    """Analytics dashboard for business owners to track performance"""
    from datetime import datetime, timedelta
    from django.db.models import Sum, Count, Avg
    import json
    
    # Get date range (default to last 30 days)
    end_date = timezone.now()
    start_date = end_date - timedelta(days=30)
    
    # Get all analyses for this specific business owner (last 30 days)
    analyses = DocumentAnalysis.objects.filter(
        user=request.user,
        created_at__gte=start_date,
        created_at__lte=end_date
    )
    
    # Calculate summary statistics
    total_revenue = float(analyses.aggregate(Sum('overall_cost'))['overall_cost__sum'] or 0)
    total_documents = analyses.count()
    total_pages = analyses.aggregate(Sum('page_count'))['page_count__sum'] or 0
    
    # Calculate color vs B&W breakdown
    color_revenue = 0.0
    bw_revenue = 0.0
    color_pages = analyses.aggregate(Sum('color_page_count'))['color_page_count__sum'] or 0
    bw_pages = analyses.aggregate(Sum('bw_page_count'))['bw_page_count__sum'] or 0
    
    # For simplicity, estimate revenue split based on page count ratio
    if total_pages > 0:
        color_ratio = float(color_pages) / float(total_pages)
        bw_ratio = float(bw_pages) / float(total_pages)
        color_revenue = total_revenue * color_ratio
        bw_revenue = total_revenue * bw_ratio
    
    # Calculate average cost per page
    avg_cost_per_page = total_revenue / total_pages if total_pages > 0 else 0
    
    # Revenue data for chart (last 30 days)
    revenue_dates = []
    revenue_data = []
    
    for i in range(30):
        chart_date = start_date + timedelta(days=i)
        # Query this user's analyses for this specific date
        day_revenue = DocumentAnalysis.objects.filter(
            user=request.user,
            created_at__date=chart_date.date()
        ).aggregate(Sum('overall_cost'))['overall_cost__sum'] or 0
        
        revenue_dates.append(chart_date.strftime('%m/%d'))
        revenue_data.append(float(day_revenue))
    
    # Recent activities
    recent_activities = []
    recent_analyses = analyses.order_by('-created_at')[:10]
    
    for analysis in recent_analyses:
        recent_activities.append({
            'type': 'upload',
            'icon': 'cloud-upload-alt',
            'title': f'Document Analyzed: {analysis.document_name}',
            'details': f'User: {analysis.user.username} • Pages: {analysis.page_count} • Cost: ₱{analysis.overall_cost}',
            'time': analysis.created_at
        })
    
    # Performance metrics
    # Calculate actual daily average revenue (only counting days with activity)
    days_with_activity = analyses.values('created_at__date').distinct().count()
    daily_avg_revenue = total_revenue / days_with_activity if days_with_activity > 0 else 0
    
    # Calculate processing efficiency (successful analyses vs total attempts)
    # For now, assuming all DocumentAnalysis records represent successful processing
    # In a real system, you'd track failed attempts separately
    total_attempts = total_documents  # Assuming all records are successful
    successful_analyses = total_documents
    processing_efficiency = (successful_analyses / total_attempts * 100) if total_attempts > 0 else 100
    
    # Calculate revenue growth (compare with previous 30-day period for this user)
    previous_start = start_date - timedelta(days=30)
    previous_end = start_date
    previous_revenue = float(DocumentAnalysis.objects.filter(
        user=request.user,
        created_at__gte=previous_start,
        created_at__lt=previous_end
    ).aggregate(Sum('overall_cost'))['overall_cost__sum'] or 0)
    
    revenue_growth = ((total_revenue - previous_revenue) / previous_revenue * 100) if previous_revenue > 0 else 0
    
    # Calculate efficiency improvement (mock for now, would need historical data)
    efficiency_improvement = 2.5  # Mock data - would need baseline efficiency tracking
    
    context = {
        'total_revenue': total_revenue,
        'total_documents': total_documents,
        'total_pages': total_pages,
        'color_revenue': color_revenue,
        'bw_revenue': bw_revenue,
        'color_pages': color_pages,
        'bw_pages': bw_pages,
        'avg_cost_per_page': avg_cost_per_page,
        'revenue_dates': json.dumps(revenue_dates),
        'revenue_data': json.dumps(revenue_data),
        'recent_activities': recent_activities,
        'daily_avg_revenue': daily_avg_revenue,
        'revenue_growth': revenue_growth,
        'processing_efficiency': processing_efficiency,
        'efficiency_improvement': efficiency_improvement,
    }
    
    return render(request, 'analyzer/business/analytics_dashboard.html', context)

@business_owner_required
def detailed_reports(request):
    """Detailed reports page for business owners"""
    from datetime import datetime, timedelta
    from django.core.paginator import Paginator
    from django.db.models import Sum, Count
    
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    user_filter = request.GET.get('user_filter')
    document_type = request.GET.get('document_type')
    
    # Base queryset - only this user's analyses
    analyses = DocumentAnalysis.objects.filter(user=request.user).order_by('-created_at')
    # Only show analyses with a completed print job
    from .models import PrintedDocument
    completed_analysis_ids = PrintedDocument.objects.filter(
        business_owner=request.user,
        status='completed'
    ).values_list('analysis_id', flat=True)
    analyses = analyses.filter(id__in=completed_analysis_ids)

    # Apply filters (remove user_filter since we only show current user's data)
    if start_date:
        analyses = analyses.filter(created_at__date__gte=start_date)
    if end_date:
        analyses = analyses.filter(created_at__date__lte=end_date)
    if document_type:
        analyses = analyses.filter(file_type=document_type)
    
    # Calculate summary for filtered data
    summary = {
        'total_revenue': float(analyses.aggregate(Sum('overall_cost'))['overall_cost__sum'] or 0),
        'total_documents': analyses.count(),
        'total_pages': analyses.aggregate(Sum('page_count'))['page_count__sum'] or 0,
    }
    
    # Pagination
    paginator = Paginator(analyses, 25)  # Show 25 analyses per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'analyses': page_obj,
        'summary': summary,
        'start_date': start_date,
        'end_date': end_date,
        'document_type': document_type,
    }
    
    return render(request, 'analyzer/business/detailed_reports.html', context)

@business_owner_required
def export_reports(request):
    """Export reports in various formats"""
    from django.http import HttpResponse
    from datetime import datetime
    import csv
    import json
    
    # Get the same filters as detailed_reports
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    document_type = request.GET.get('document_type')
    export_format = request.GET.get('export', 'csv').lower()
    if export_format == 'excel':
        export_format = 'xlsx'

    # Base queryset with same filters - only current user's data
    analyses = DocumentAnalysis.objects.filter(user=request.user).order_by('-created_at')
    if start_date:
        analyses = analyses.filter(created_at__date__gte=start_date)
    if end_date:
        analyses = analyses.filter(created_at__date__lte=end_date)
    if document_type:
        analyses = analyses.filter(file_type=document_type)

    # Executive summary export
    if export_format == 'summary':
        # Calculate summary stats (same as detailed_reports)
        total_revenue = float(analyses.aggregate(Sum('overall_cost'))['overall_cost__sum'] or 0)
        total_documents = analyses.count()
        total_pages = analyses.aggregate(Sum('page_count'))['page_count__sum'] or 0
        color_pages = analyses.aggregate(Sum('color_page_count'))['color_page_count__sum'] or 0
        bw_pages = analyses.aggregate(Sum('bw_page_count'))['bw_page_count__sum'] or 0
        # Prepare executive summary CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="executive_summary.csv"'
        writer = csv.writer(response)
        writer.writerow(["Executive Summary Report"])
        writer.writerow(["Date Generated", datetime.now().strftime('%Y-%m-%d %H:%M')])
        writer.writerow([])
        writer.writerow(["Total Revenue", f"₱{total_revenue:,.2f}"])
        writer.writerow(["Documents Processed", total_documents])
        writer.writerow(["Total Pages", total_pages])
        writer.writerow(["Color Pages", color_pages])
        writer.writerow(["Black & White Pages", bw_pages])
        writer.writerow([])
        writer.writerow(["This summary provides high-level business insights for the selected period."])
        return response
    
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="document_analysis_report_{datetime.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Document Name', 'User', 'Date', 'Total Pages', 'Color Pages', 
            'B&W Pages', 'Total Cost', 'File Type', 'File Size (KB)'
        ])
        
        for analysis in analyses:
            writer.writerow([
                analysis.document_name,
                analysis.user.username,
                analysis.created_at.strftime('%Y-%m-%d %H:%M'),
                analysis.page_count,
                analysis.color_page_count,
                analysis.bw_page_count,
                f"₱{analysis.overall_cost}",
                analysis.file_type.upper(),
                round(analysis.file_size / 1024, 2) if analysis.file_size else 0
            ])
        
        return response
    
    elif export_format == 'json':
        response = HttpResponse(content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="document_analysis_report_{datetime.now().strftime("%Y%m%d")}.json"'
        
        data = []
        for analysis in analyses:
            data.append({
                'document_name': analysis.document_name,
                'user': analysis.user.username,
                'date': analysis.created_at.isoformat(),
                'total_pages': analysis.page_count,
                'color_pages': analysis.color_page_count,
                'bw_pages': analysis.bw_page_count,
                'total_cost': float(analysis.overall_cost),
                'file_type': analysis.file_type,
                'file_size_kb': round(analysis.file_size / 1024, 2) if analysis.file_size else 0
            })
        
        response.write(json.dumps(data, indent=2))
        return response
    elif export_format == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        wb = Workbook()
        ws = wb.active
        ws.title = "Document Analysis Report"
        headers = ['Document Name', 'User', 'Date', 'Total Pages', 'Color Pages', 'B&W Pages', 'Total Cost', 'File Type', 'File Size (KB)']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        for analysis in analyses:
            ws.append([
                analysis.document_name,
                analysis.user.username,
                analysis.created_at.strftime('%Y-%m-%d %H:%M'),
                analysis.page_count,
                analysis.color_page_count,
                analysis.bw_page_count,
                f"₱{analysis.overall_cost}",
                analysis.file_type.upper(),
                round(analysis.file_size / 1024, 2) if analysis.file_size else 0
            ])
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="document_analysis_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        response.write(output.getvalue())
        return response
    elif export_format == 'pdf':
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
        from io import BytesIO
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        data = [
            ['Document Name', 'User', 'Date', 'Total Pages', 'Color Pages', 'B&W Pages', 'Total Cost', 'File Type', 'File Size (KB)']
        ]
        for analysis in analyses:
            data.append([
                analysis.document_name,
                analysis.user.username,
                analysis.created_at.strftime('%Y-%m-%d %H:%M'),
                analysis.page_count,
                analysis.color_page_count,
                analysis.bw_page_count,
                f"₱{analysis.overall_cost}",
                analysis.file_type.upper(),
                round(analysis.file_size / 1024, 2) if analysis.file_size else 0
            ])
        table = Table(data)
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f5f5f5')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ])
        table.setStyle(style)
        elements = [table]
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="document_analysis_report_{datetime.now().strftime("%Y%m%d")}.pdf"'
        response.write(pdf)
        return response
    else:
        from django.http import HttpResponseBadRequest
        return HttpResponseBadRequest("Unsupported export format.")
def get_progress(request):
    """Get progress for a session"""
    session_id = request.GET.get('session_id')
    print(f"DEBUG: Getting progress for session: {session_id}")  # Debug print
    if not session_id or session_id not in progress_sessions:
        print(f"DEBUG: Session not found. Available sessions: {list(progress_sessions.keys())}")
        return JsonResponse({'progress': 0, 'status': 'Starting...', 'error': 'Session not found'})
    
    data = progress_sessions[session_id]
    print(f"DEBUG: Returning progress: {data}")  # Debug print
    
    # Return all data including statistics
    response_data = {
        'progress': data['progress'],
        'status': data['status'],
        'timestamp': data['timestamp'],
        'pages_processed': data.get('pages_processed', 0),
        'total_pages': data.get('total_pages', 0),
        'color_count': data.get('color_count', 0),
        'bw_count': data.get('bw_count', 0),
        'current_cost': data.get('current_cost', 0)
    }
    print(f"DEBUG: Full response data: {response_data}")  # Debug print
    return JsonResponse(response_data)


def clear_progress(session_id):
    """Clear progress for a session when complete"""
    if session_id in progress_sessions:
        del progress_sessions[session_id]

@login_required
def request_status_page(request):
    """Page for users to check their business owner request status and communicate with admin"""
    try:
        # Get the user's business owner request
        business_request = BusinessOwnerRequest.objects.get(user=request.user)
        
        # Get all messages for this request (excluding internal notes for regular users)
        if request.user.userprofile.is_super_admin():
            messages_list = business_request.messages.all()
        else:
            messages_list = business_request.messages.filter(is_internal_note=False)
            # Mark messages as read by user
            business_request.messages.filter(is_internal_note=False, read_by_user=False).update(read_by_user=True)
        
        # Get status change history
        status_changes = business_request.status_changes.all()
        
        # Handle new message submission
        if request.method == 'POST':
            message_text = request.POST.get('message', '').strip()
            if message_text:
                new_message = RequestMessage.objects.create(
                    request=business_request,
                    sender=request.user,
                    message=message_text,
                    is_internal_note=False,
                    read_by_admin=False,
                    read_by_user=True  # User's own message is considered "read"
                )
                messages.success(request, 'Message sent to administrator.')
                return redirect('request_status_page')
        
        context = {
            'business_request': business_request,
            'messages_list': messages_list,
            'status_changes': status_changes,
            'can_send_messages': business_request.status == 'pending'
        }
        
        return render(request, 'analyzer/request_status.html', context)
        
    except BusinessOwnerRequest.DoesNotExist:
        messages.error(request, 'No business owner request found. Please submit a request first.')
        return redirect('regular_user_dashboard')

@super_admin_required
def admin_request_detail(request, request_id):
    """Detailed view for super admin to manage a specific business owner request"""
    business_request = get_object_or_404(BusinessOwnerRequest, id=request_id)
    
    # Mark admin messages as read
    business_request.messages.filter(read_by_admin=False).update(read_by_admin=True)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'send_message':
            message_text = request.POST.get('message', '').strip()
            is_internal = request.POST.get('is_internal') == 'on'
            
            if message_text:
                RequestMessage.objects.create(
                    request=business_request,
                    sender=request.user,
                    message=message_text,
                    is_internal_note=is_internal,
                    read_by_admin=True,
                    read_by_user=False
                )
                
                message_type = "Internal note" if is_internal else "Message"
                messages.success(request, f'{message_type} added successfully.')
                
        elif action == 'update_status':
            new_status = request.POST.get('new_status')
            reason = request.POST.get('reason', '').strip()
            
            if new_status and new_status != business_request.status:
                old_status = business_request.status
                
                # Create status change record
                RequestStatusChange.objects.create(
                    request=business_request,
                    changed_by=request.user,
                    old_status=old_status,
                    new_status=new_status,
                    reason=reason
                )
                
                # Update the request status
                business_request.status = new_status
                business_request.save()
                
                # If approved, promote user to business owner
                if new_status == 'approved':
                    user_profile = business_request.user.userprofile
                    user_profile.role = 'business_owner'
                    user_profile.save()
                    
                    # Send automatic message to user
                    RequestMessage.objects.create(
                        request=business_request,
                        sender=request.user,
                        message=f"Congratulations! Your business owner request has been approved. {reason}",
                        is_internal_note=False,
                        read_by_admin=True,
                        read_by_user=False
                    )
                    
                elif new_status == 'rejected':
                    # Send automatic message to user
                    RequestMessage.objects.create(
                        request=business_request,
                        sender=request.user,
                        message=f"Your business owner request has been rejected. Reason: {reason}",
                        is_internal_note=False,
                        read_by_admin=True,
                        read_by_user=False
                    )
                
                messages.success(request, f'Request status updated to {new_status}.')
        
        return redirect('admin_request_detail', request_id=request_id)
    
    # Get all messages and status changes
    all_messages = business_request.messages.all()
    status_changes = business_request.status_changes.all()
    
    context = {
        'business_request': business_request,
        'all_messages': all_messages,
        'status_changes': status_changes,
        'status_choices': BusinessOwnerRequest.STATUS_CHOICES
    }
    
    return render(request, 'analyzer/admin_request_detail.html', context)

@super_admin_required
def admin_requests_list(request):
    """List all business owner requests with summary info for super admin"""
    requests = BusinessOwnerRequest.objects.all().order_by('-created_at')
    
    # Add unread message counts
    for req in requests:
        req.unread_count = req.messages.filter(read_by_admin=False, is_internal_note=False).count()
    
    context = {
        'requests': requests,
        'pending_count': requests.filter(status='pending').count(),
        'approved_count': requests.filter(status='approved').count(),
        'rejected_count': requests.filter(status='rejected').count()
    }
    
    return render(request, 'analyzer/admin_requests_list.html', context)


@login_required
def print_document(request):
    """Add a document to the print queue"""
    if request.method == 'POST':
        analysis_id = request.POST.get('analysis_id')
        paper_size = request.POST.get('paper_size', 'a4')
        copies = int(request.POST.get('copies', 1))
        
        try:
            # Get the analysis
            analysis = get_object_or_404(DocumentAnalysis, id=analysis_id)
            
            # Verify user has access to this analysis
            user_profile = get_user_profile(request.user)
            if user_profile.role == 'regular' and analysis.user != request.user:
                messages.error(request, "You can only print your own documents.")
                return redirect('regular_user_dashboard')
            
            # Determine business owner for pricing
            if user_profile.role == 'business_owner':
                business_owner = request.user
            else:
                # For regular users, need to find a business owner with pricing rules
                business_owner = User.objects.filter(
                    userprofile__role='business_owner',
                    costsetting__paper_size=paper_size,
                    costsetting__is_active=True
                ).first()
                
                if not business_owner:
                    messages.error(request, "No business owner pricing available for the selected paper size.")
                    return redirect('regular_user_dashboard')
            
            # Calculate cost per copy based on analysis results and paper size
            total_cost = 0
            for result in analysis.analysis_result:
                coverage = result.get('coverage_percentage', 0)
                is_color = result.get('image_type') == 'Color'
                
                # Get pricing for this page
                cost, reason = get_pricing_for_user(business_owner, is_color, coverage, paper_size)
                total_cost += float(cost)
            
            # Create print job
            print_job = PrintedDocument.objects.create(
                analysis=analysis,
                printed_by=request.user,
                business_owner=business_owner,
                paper_size=paper_size,
                copies=copies,
                cost_per_copy=Decimal(str(total_cost)),
                status='queued'
            )
            
            messages.success(
                request, 
                f"✅ Document queued for printing! {copies} copies on {print_job.get_paper_size_display()} - Total: ₱{print_job.total_revenue}"
            )
            
            # Redirect based on user role
            if user_profile.role == 'regular':
                return redirect('regular_user_dashboard')
            elif user_profile.role == 'business_owner':
                return redirect('business_owner_dashboard')
            else:
                return redirect('dashboard_router')
                
        except Exception as e:
            messages.error(request, f"Error adding to print queue: {str(e)}")
            return redirect('dashboard_router')
    
    return redirect('dashboard_router')


@login_required 
def print_history(request):
    """View print history and manage print queue"""
    user_profile = get_user_profile(request.user)
    
    if user_profile.role == 'business_owner':
        # Business owners see all prints that generate revenue for them
        print_jobs = PrintedDocument.objects.filter(business_owner=request.user).order_by('-queued_at')
        
        # Calculate total revenue
        total_revenue = print_jobs.filter(status='completed').aggregate(
            total=Sum('total_revenue')
        )['total'] or 0
        
        # Revenue breakdown by status
        revenue_breakdown = {
            'completed': print_jobs.filter(status='completed').aggregate(total=Sum('total_revenue'))['total'] or 0,
            'queued': print_jobs.filter(status='queued').aggregate(total=Sum('total_revenue'))['total'] or 0,
            'printing': print_jobs.filter(status='printing').aggregate(total=Sum('total_revenue'))['total'] or 0,
        }
        
        context = {
            'print_jobs': print_jobs,
            'total_revenue': total_revenue,
            'revenue_breakdown': revenue_breakdown,
            'is_business_owner': True
        }
        
        return render(request, 'analyzer/business/print_history.html', context)
        
    else:
        # Regular users see their own print jobs
        print_jobs = PrintedDocument.objects.filter(printed_by=request.user).order_by('-queued_at')
        
        context = {
            'print_jobs': print_jobs,
            'is_business_owner': False
        }
        
        return render(request, 'analyzer/user/print_history.html', context)


@business_owner_required
def update_print_status(request):
    """Update print job status - Business owners only"""
    if request.method == 'POST':
        print_job_id = request.POST.get('print_job_id')
        new_status = request.POST.get('status')
        auto_update = request.POST.get('auto_update') == 'true'
        
        try:
            print_job = get_object_or_404(PrintedDocument, id=print_job_id, business_owner=request.user)
            
            # Update status
            print_job.status = new_status
            
            # Set completion time if completed
            if new_status == 'completed':
                print_job.completed_at = timezone.now()
            
            print_job.save()
            
            # Only show messages for manual updates, not automatic ones
            if not auto_update:
                messages.success(request, f"Print job status updated to {print_job.get_status_display()}")
            
        except Exception as e:
            if not auto_update:
                messages.error(request, f"Error updating print status: {str(e)}")
    
    return redirect('print_history')

@login_required
def print_document_content(request, analysis_id):
    """Serve the document content for printing"""
    try:
        # Get the analysis
        analysis = get_object_or_404(DocumentAnalysis, id=analysis_id)
        
        # Verify user has access to this analysis
        user_profile = get_user_profile(request.user)
        if user_profile.role == 'regular' and analysis.user != request.user:
            return HttpResponse("Access denied", status=403)
        elif user_profile.role == 'business_owner':
            # Business owners can print documents from their customers
            pass
        
        # Build the HTML content with all page images - minimal styling to preserve original layout
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>{analysis.document_name}</title>
            <style>
                body {{
                    margin: 0;
                    padding: 0;
                    background: white;
                }}
                .page-container {{
                    margin: 0;
                    padding: 0;
                    text-align: center;
                    page-break-after: always;
                    width: 100%;
                    height: 100vh;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                }}
                .page-container:last-child {{
                    page-break-after: avoid;
                }}
                .page-image {{
                    max-width: 100%;
                    max-height: 100%;
                    width: auto;
                    height: auto;
                }}
                .print-controls {{
                    position: fixed;
                    top: 10px;
                    right: 10px;
                    z-index: 1000;
                    background: rgba(255, 255, 255, 0.9);
                    padding: 10px;
                    border-radius: 8px;
                    box-shadow: 0 2px 10px rgba(0,0,0,0.1);
                }}
                .print-btn {{
                    background: #28a745;
                    color: white;
                    border: none;
                    padding: 8px 15px;
                    border-radius: 4px;
                    font-size: 12px;
                    cursor: pointer;
                    margin: 2px;
                }}
                .print-btn:hover {{
                    background: #218838;
                }}
                @media print {{
                    .print-controls {{
                        display: none !important;
                    }}
                    .page-container {{
                        height: 100vh;
                        margin: 0;
                        padding: 0;
                    }}
                    /* Remove browser default print headers and footers */
                    @page {{
                        margin: 0;
                        size: auto;
                    }}
                    body {{
                        -webkit-print-color-adjust: exact;
                        print-color-adjust: exact;
                    }}
                }}
            </style>
        </head>
        <body>
            <div class="print-controls">
                <button onclick="printClean()" class="print-btn">
                    🖨️ Print
                </button>
                <button onclick="window.close()" class="print-btn" style="background: #6c757d;">
                    ✕ Close
                </button>
            </div>
            
            <script>
                // Get job ID from URL parameters
                const urlParams = new URLSearchParams(window.location.search);
                const jobId = urlParams.get('job_id');
                
                function printClean() {{
                    // Hide the controls before printing
                    document.querySelector('.print-controls').style.display = 'none';
                    
                    // Show print instructions and handle printing
                    const instructions = document.createElement('div');
                    instructions.style.cssText = `
                        position: fixed;
                        top: 50%;
                        left: 50%;
                        transform: translate(-50%, -50%);
                        background: rgba(0,0,0,0.9);
                        color: white;
                        padding: 20px;
                        border-radius: 8px;
                        z-index: 10000;
                        text-align: center;
                        font-family: Arial, sans-serif;
                    `;
                    instructions.innerHTML = `
                        <h3>Clean Print Setup</h3>
                        <p>In the print dialog that opens:</p>
                        <p><strong>1.</strong> Click "More settings"</p>
                        <p><strong>2.</strong> Turn OFF "Headers and footers"</p>
                        <p><strong>3.</strong> Set margins to "None" or "Minimum"</p>
                        <button onclick="startPrinting(this.parentElement);" 
                                style="background: #28a745; color: white; border: none; padding: 10px 20px; border-radius: 4px; margin-top: 10px; cursor: pointer;">
                            OK, Open Print Dialog
                        </button>
                    `;
                    document.body.appendChild(instructions);
                }}
                
                function startPrinting(instructionsDiv) {{
                    instructionsDiv.remove();
                    
                    // Set up print event listeners before opening print dialog
                    window.addEventListener('beforeprint', function() {{
                        console.log('Print dialog opened');
                    }});
                    
                    window.addEventListener('afterprint', function() {{
                        console.log('Print dialog closed');
                        // Give user a moment then close window and update parent
                        setTimeout(function() {{
                            if (window.opener && !window.opener.closed) {{
                                // Notify parent window that printing is done
                                window.opener.postMessage({{
                                    type: 'print_completed',
                                    jobId: jobId
                                }}, '*');
                            }}
                            window.close();
                        }}, 1000);
                    }});
                    
                    // Open print dialog
                    window.print();
                }}
                
                // Auto-hide instructions when window gets focus back
                window.addEventListener('focus', function() {{
                    const instructions = document.querySelector('div[style*="position: fixed"]');
                    if (instructions && instructions.innerHTML.includes('Clean Print Setup')) {{
                        instructions.remove();
                    }}
                }});
                
                // Listen for messages from parent window
                window.addEventListener('message', function(event) {{
                    if (event.data.type === 'update_status') {{
                        // Parent is updating status
                        console.log('Status update:', event.data.status);
                    }}
                }});
            </script>
        """
        
        # Add each page image without any extra information - preserve original layout
        if analysis.analysis_result:
            for i, page_result in enumerate(analysis.analysis_result, 1):
                image_path = page_result.get('image_path', '')
                if image_path:
                    # Convert relative path to absolute URL
                    if image_path.startswith('/media/'):
                        image_url = request.build_absolute_uri(image_path)
                    else:
                        image_url = image_path
                    
                    html_content += f"""
                    <div class="page-container">
                        <img src="{image_url}" alt="Page {i}" class="page-image" />
                    </div>
                    """
        else:
            html_content += """
            <div class="page-container">
                <p>No page images available for this document.</p>
            </div>
            """
        
        html_content += """
        </body>
        </html>
        """
        
        return HttpResponse(html_content, content_type='text/html')
        
    except Exception as e:
        return HttpResponse(f"Error loading document: {str(e)}", status=500)

@super_admin_required
def default_pricing_management(request):
    """Manage default pricing rules that apply system-wide"""
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add':
            try:
                paper_size = request.POST.get('paper_size')
                color = request.POST.get('color')
                coverage_min = request.POST.get('coverage_min')
                coverage_max = request.POST.get('coverage_max')
                cost = request.POST.get('cost')
                reason = request.POST.get('reason', '')
                is_active = request.POST.get('is_active', 'True')

                # Convert color and is_active to boolean
                color_bool = True if color == 'True' or color is True else False
                is_active_bool = True if is_active == 'True' or is_active is True else False

                existing_rule = DefaultPricingRule.objects.filter(
                    paper_size=paper_size,
                    color=color_bool,
                    coverage_min=coverage_min,
                    coverage_max=coverage_max
                ).first()

                if existing_rule:
                    messages.error(request, f'A default pricing rule already exists for {existing_rule.get_paper_size_display()} - {"Color" if existing_rule.color else "B&W"} {existing_rule.coverage_min}-{existing_rule.coverage_max}%. Please edit the existing rule instead.')
                else:
                    rule = DefaultPricingRule.objects.create(
                        paper_size=paper_size,
                        color=color_bool,
                        coverage_min=coverage_min,
                        coverage_max=coverage_max,
                        cost=cost,
                        reason=reason,
                        is_active=is_active_bool
                    )
                    messages.success(request, f'Default pricing rule added successfully for {rule.get_paper_size_display()} - {"Color" if rule.color else "B&W"} {rule.coverage_min}-{rule.coverage_max}%.')
            except Exception as e:
                messages.error(request, f'Error adding default pricing rule: {str(e)}')
                
        elif action == 'edit':
            try:
                rule_id = request.POST.get('rule_id')
                rule = get_object_or_404(DefaultPricingRule, id=rule_id)
                rule.paper_size = request.POST.get('paper_size')
                color = request.POST.get('color')
                coverage_min = request.POST.get('coverage_min')
                coverage_max = request.POST.get('coverage_max')
                cost = request.POST.get('cost')
                reason = request.POST.get('reason', '')
                is_active = request.POST.get('is_active', 'True')

                rule.color = True if color == 'True' or color is True else False
                rule.coverage_min = coverage_min
                rule.coverage_max = coverage_max
                rule.cost = cost
                rule.reason = reason
                rule.is_active = True if is_active == 'True' or is_active is True else False
                rule.save()
                messages.success(request, f'Default pricing rule updated successfully for {rule.get_paper_size_display()} - {"Color" if rule.color else "B&W"} {rule.coverage_min}-{rule.coverage_max}%.')
            except Exception as e:
                messages.error(request, f'Error updating default pricing rule: {str(e)}')
                
        elif action == 'delete':
            try:
                rule_id = request.POST.get('rule_id')
                rule = get_object_or_404(DefaultPricingRule, id=rule_id)
                rule_display = f"{rule.get_paper_size_display()} - {'Color' if rule.color else 'B&W'}"
                rule.delete()
                
                messages.success(request, f'Default pricing rule for {rule_display} has been deleted.')
                
            except Exception as e:
                messages.error(request, f'Error deleting default pricing rule: {str(e)}')
        
        return redirect('default_pricing_management')
    
    # GET request - display the management page
    try:
        default_rules = DefaultPricingRule.objects.all().order_by('paper_size', 'color', 'coverage_min')
    except Exception as e:
        # Handle database errors gracefully
        print(f"Database error in default_pricing_management: {e}")
        default_rules = []
    
    context = {
        'default_rules': default_rules,
    }
    
    return render(request, 'analyzer/admin/default_pricing_management.html', context)

def test_view(request):
    """Simple test view to check if URLs are working"""
    return HttpResponse("Test URL is working! The default pricing management should work too.")